import os
import sys
from datetime import datetime
import pytz
from dateutil import parser

import pandas as pd
import requests

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side

from config.server.server_config import server_config
from config.system.log_config import setup_logging, set_current_server
from managers.dell_server_manager import DellServerManager
from PyQt6.QtCore import QTimer, Qt, QUrl
from PyQt6.QtGui import QDesktopServices
from PyQt6.QtWidgets import (QApplication, QDialog, QFileDialog, QGroupBox, QHBoxLayout, QLabel, QLineEdit, 
                             QMainWindow, QMessageBox, QPushButton, QProgressDialog, QVBoxLayout, QWidget)
from ui.components.popups.detail_dialog import DetailDialog
from ui.components.popups.error_dialog import ErrorDialog
from ui.components.server_section import ServerSection
from utils.utils import convert_capacity

logger = setup_logging()

class SystemInfoGroup(QGroupBox):
    def __init__(self):
        super().__init__("시스템 정보")
        self._init_ui()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        self.labels = {
            'model': QLabel("모델명: 연결 대기 중..."),
            'service_tag': QLabel("서비스태그: 연결 대기 중..."),
            'bios': QLabel("BIOS: 연결 대기 중..."),
            'idrac': QLabel("iDRAC: 연결 대기 중..."),
        }
        for label in self.labels.values():
            label.setTextInteractionFlags(
                Qt.TextInteractionFlag.TextSelectableByMouse | 
                Qt.TextInteractionFlag.TextSelectableByKeyboard |
                Qt.TextInteractionFlag.LinksAccessibleByMouse
            )
            label.setCursor(Qt.CursorShape.IBeamCursor)
            layout.addWidget(label)
    
    def update_info(self, **info):
        try:
            mappings = {
                'model': ('모델명', info.get('model', '알 수 없음')),
                'service_tag': ('서비스태그', info.get('service_tag', '알 수 없음')),
                'bios': ('BIOS', info.get('bios_version', '알 수 없음')),
                'idrac': ('iDRAC', info.get('idrac_version', '알 수 없음')),
            }
            for key, (label, value) in mappings.items():
                self.labels[key].setText(f"{label}: {value}")
                
        except Exception as e:
            logger.error(f"Error updating system info: {str(e)}")
            error_dialog = ErrorDialog(
                "오류 발생",
                "시스템 정보를 업데이트하는 중 오류가 발생했습니다.",
                str(e)
            )
            error_dialog.exec()
    
    def set_loading_state(self):
        for key in self.labels:
            self.labels[key].setText(f"{key.replace('_', ' ').title()}: 로딩 중...")
    
    def set_error_state(self):
        for key in self.labels:
            self.labels[key].setText(f"{key.replace('_', ' ').title()}: 연결 실패")

class HardwareInfoWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.logger = setup_logging()  # logger 추가
        self.main_layout = QHBoxLayout(self)
        self.main_layout.setSpacing(5)
        self.server_manager = None
        self.parent_window = parent  # 부모 윈도우 저장
        self.setup_ui()
        self._setup_connections(parent)

    def setup_ui(self):
        # 시스템 정보
        self.system_info = SystemInfoGroup()
        self.main_layout.addWidget(self.system_info)
        
        # 상태 정보
        self.status_info = QGroupBox("상태 정보")
        status_layout = QVBoxLayout(self.status_info)
        self.status_labels = {}
        for item in ["CPU", "MEM", "DSK", "PWR"]:
            label = QLabel(f"{item}: 연결 대기 중...")
            label.setTextInteractionFlags(Qt.TextInteractionFlag.LinksAccessibleByMouse)
            label.setCursor(Qt.CursorShape.PointingHandCursor)
            label.mousePressEvent = lambda e, item=item: self.show_component_details(item)
                    
            status_layout.addWidget(label)
            self.status_labels[item] = label
        self.main_layout.addWidget(self.status_info)
        
        # 편의 기능
        self.quick_action = QGroupBox("편의 기능")
        quick_layout = QVBoxLayout(self.quick_action)
        actions = ["🔗 빠른 연결 실행", "🔄 재시작", "💾 저장", "🔍 지원"]
        
        for action in actions:
            btn = QPushButton(action)
            btn.setFixedHeight(25)
            if action == "🔗 빠른 연결 실행":
                btn.clicked.connect(self._on_quick_connect)
            elif action == "🔄 재시작":
                btn.clicked.connect(self.restart_application)
            elif action == "💾 저장":
                btn.clicked.connect(lambda: save_system_info(self, self.server_manager))
            elif action == "🔍 지원":
                btn.clicked.connect(self.open_dell_support)
            quick_layout.addWidget(btn)
        self.main_layout.addWidget(self.quick_action)

    def open_dell_support(self):
        try:
            model_name = ''
            service_tag = ''
            
            # 서버가 연결되어 있는 경우에만 정보 가져오기
            if self.server_manager is not None:
                try:
                    basic_info = self.server_manager.fetch_basic_info()
                    system_info = basic_info['system']
                    bios_info = basic_info['bios']
                    
                    model_name = bios_info['Attributes'].get('SystemModelName', '').lower()
                    service_tag = system_info.get('ServiceTag', '')
                except Exception as e:
                    logger.error(f"서버 정보 가져오기 실패: {str(e)}")
            
            # 팝업 다이얼로그 생성
            dialog = QDialog(self)
            dialog.setWindowTitle("Dell 지원 사이트")
            dialog.setFixedWidth(400)

            layout = QVBoxLayout()
            layout.addSpacing(10)

            # 서버 연결 상태에 따라 다른 메시지 표시
            if self.server_manager is not None:
                info_label = QLabel("Dell 지원 사이트에 접속하실 방법을 선택해주세요:")
            else:
                info_label = QLabel("서버가 연결되어 있지 않습니다.\n아래에서 직접 검색이 가능합니다.")
                info_label.setStyleSheet("color: #666;")
            info_label.setWordWrap(True)
            layout.addWidget(info_label)

            # 서버가 연결된 경우에만 모델/서비스태그 버튼 표시
            if self.server_manager is not None:
                # 모델 기반 버튼
                model_url = self.get_dell_url(model=model_name)
                if model_url:
                    model_btn = QPushButton(f"{model_name.upper()} 모델 페이지로 이동")
                    model_btn.clicked.connect(lambda: self.open_url(model_url))
                    layout.addWidget(model_btn)

                # 서비스 태그 기반 버튼
                if service_tag:
                    tag_btn = QPushButton(f"서비스 태그({service_tag}) 페이지로 이동")
                    tag_btn.clicked.connect(lambda: self.open_url(self.get_dell_url(tag=service_tag)))
                    layout.addWidget(tag_btn)

            # 직접 검색 섹션
            search_group = QGroupBox("직접 검색")
            search_layout = QVBoxLayout()

            # 검색 도움말 추가
            search_help = QLabel(
                "• 서비스 태그 검색: 7자리 영문자+숫자 조합 (예: 9QVLH04)\n"
                "• 모델명 검색: PowerEdge 서버 모델명 (예: R650, R750)"
            )
            search_help.setStyleSheet("color: gray;")
            search_help.setWordWrap(True)

            search_input = QLineEdit()
            search_input.setPlaceholderText("모델명 또는 서비스 태그 입력")
            search_btn = QPushButton("검색")

            def search_dell_support():
                try:
                    search_term = search_input.text().strip()
                    if not search_term:
                        QMessageBox.warning(dialog, "입력 오류", "검색어를 입력해주세요.")
                        return

                    # 서비스 태그 형식 검사 (7자리 영문자+숫자)
                    if len(search_term) == 7 and search_term.isalnum():
                        url = self.get_dell_url(tag=search_term)
                    else:
                        # 모델명으로 처리
                        url = self.get_dell_url(model=search_term)

                    if url:
                        self.open_url(url)
                        dialog.accept()
                    else:
                        QMessageBox.warning(
                            dialog,
                            "검색 오류",
                            "올바른 모델명 또는 서비스 태그를 입력해주세요."
                        )
                except Exception as e:
                    QMessageBox.warning(dialog, "검색 오류", str(e))

            search_btn.clicked.connect(search_dell_support)
            search_input.returnPressed.connect(search_dell_support)

            search_layout.addWidget(search_help)
            search_layout.addWidget(search_input)
            search_layout.addWidget(search_btn)
            search_group.setLayout(search_layout)
            layout.addWidget(search_group)

            dialog.setLayout(layout)
            dialog.exec()

        except Exception as e:
            logger.error(f"Dell 지원 사이트 접근 오류: {str(e)}")
            error_dialog = ErrorDialog(
                "지원 사이트 접근 오류",
                "Dell 지원 사이트 접근 중 오류가 발생했습니다.",
                str(e),
                self
            )
            error_dialog.exec()

    def open_url(self, url):
        QDesktopServices.openUrl(QUrl(url))

    def get_dell_url(self, model=None, tag=None):
        """Dell 지원 URL 생성 함수"""
        base_url = "https://www.dell.com/support/home/ko-kr/product-support/"
        if tag:
            url = f"{base_url}servicetag/{tag}/overview"
            logger.debug(f"Dell 지원 사이트 접근: 서비스 태그 기반 URL - {url}")
            return url
        elif model:
            # 모델명에서 시리즈와 숫자 분리
            model_parts = model.lower().split()
            for part in model_parts:
                if 'r' in part and any(c.isdigit() for c in part):
                    # R650, R750 등의 형식에서 'r'과 숫자 추출
                    series = 'poweredge'
                    number = ''.join(filter(str.isdigit, part))
                    formatted_model = f"{series}-r{number}"
                    url = f"{base_url}product/{formatted_model}/overview"
                    logger.debug(f"Dell 지원 사이트 접근: 모델 기반 URL - {url}")
                    return url
            
            # 지원되지 않는 모델인 경우
            logger.debug(f"지원되지 않는 모델: {model}")
            return None

    def restart_application(self):
        logger.info("프로그램 재시작 시도")
        python = sys.executable
        script = sys.argv[0]
        
        # 새 프로세스 시작
        logger.info("새 프로세스 시작")
        import subprocess
        subprocess.Popen([python, script])
        
        # 현재 창 닫기
        logger.info("프로그램 종료")
        QApplication.closeAllWindows()
    
    def _setup_connections(self, parent):
        if isinstance(parent, QMainWindow):
            server_section = parent.findChild(ServerSection)
            if server_section:
                server_section.server_connection_changed.connect(self.on_server_connected)
    
    def on_server_connected(self, server_name, connected):
        """서버 연결 상태 변경 시 호출되는 메서드"""
        try:
            if not connected:
                self.logger.warning("서버가 연결되지 않았습니다.")
                return
                
            if server_name not in server_config.servers:
                self.logger.warning(f"서버 정보를 찾을 수 없습니다: {server_name}")
                return
                
            server_info = {
                'NAME': server_name,
                'IP': server_config.servers[server_name].IP,  # 대문자로 변경
                'PORT': server_config.servers[server_name].PORT,  # 대문자로 변경
                'USERNAME': server_config.servers[server_name].USERNAME,  # 대문자로 변경
                'PASSWORD': server_config.servers[server_name].PASSWORD  # 대문자로 변경
            }
            
            self.server_manager = DellServerManager(
                ip=server_info['IP'],  # 대문자로 변경
                port=server_info['PORT'],  # 대문자로 변경
                auth=(server_info['USERNAME'], server_info['PASSWORD'])  # 대문자로 변경
            )
            
            self.system_info.set_loading_state()
            QTimer.singleShot(0, lambda: self.update_system_info())
            
        except Exception as e:
            self.logger.error(f"시스템 정보 업데이트 중 오류 발생: {str(e)}")
            if hasattr(self, 'system_info'):
                self.system_info.set_error_state()

    def update_status_labels(self, status_dict):
        """상태 레이블 업데이트"""
        for key, value in status_dict.items():
            if key in self.status_labels:
                self.status_labels[key].setText(value)

    def clear_status_info(self):
        """상태 정보 초기화"""
        for label in self.status_labels.values():
            label.setText(f"{label.objectName()}: 연결 끊김")

    def update_system_info(self):
        """서버 연결 시 시스템 정보 업데이트"""
        try:
            if not hasattr(self, 'server_manager') or not self.server_manager:
                main_window = self.window()
                if hasattr(main_window, 'server_section'):
                    self.server_manager = main_window.server_section.server_manager
                else:
                    return
            
            if not self.server_manager:
                return
                
            # 시스템 정보 업데이트
            basic_info = self.server_manager.fetch_basic_info()
            system_data = basic_info.get('system', {})
            bios_data = basic_info.get('bios', {})
            idrac_data = basic_info.get('idrac', {})
            
            if system_data:
                self.system_info.update_info(
                    model=system_data.get('Model'),
                    service_tag=system_data.get('ServiceTag'),
                    bios_version=bios_data.get('Attributes', {}).get('SystemBiosVersion'),
                    idrac_version=idrac_data.get('FirmwareVersion')
                )
                
                # 상태 정보 업데이트
                self._update_all_info()
                
        except Exception as e:
            logger.error(f"시스템 정보 업데이트 실패: {str(e)}")
            self.system_info.set_error_state()

    def _update_all_info(self):
        """모든 상태 정보 업데이트"""
        try:
            # CPU 상태 정보 업데이트
            self._update_cpu_status()

            # 메모리 상태 정보 업데이트
            self._update_memory_status()

            # physical disk 상태 정보 업데이트
            self._update_disk_status()

            # PSU 상태 정보 업데이트
            self._update_psu_status()
            
        except Exception as e:
            logger.error(f"상태 정보 업데이트 실패: {str(e)}")
            for label in ['CPU', 'MEM', 'DSK', 'PWR']:
                self.status_labels[label].setText(f"{label}: 상태 확인 실패")

            error_dialog = ErrorDialog(
                "정보 업데이트 오류",
                "시스템 정보를 업데이트하는 중 오류가 발생했습니다.",
                str(e),
                self
            )
            error_dialog.exec()

    def _update_cpu_status(self):
        try:
            if self.server_manager is not None:
                processors_data = self.server_manager.fetch_processors_info()
                if processors_data.get('Members@odata.count', 0) > 0:
                    cpu_count = {"✅": 0, "❌": 0, "⚠️": 0}
                    
                    for member in processors_data.get('Members', []):
                        member_uri = member.get('@odata.id')
                        if member_uri:
                            cpu_response = requests.get(
                                f"{self.server_manager.endpoints.base_url}{member_uri}",
                                auth=self.server_manager.auth,
                                verify=False
                            )
                            cpu_info = cpu_response.json()
                            enabled = cpu_info.get('Enabled', True)
                            status = cpu_info.get('Status', {})
                            health = status.get('Health', 'Unknown')
                            
                            if not enabled:
                                cpu_count["❌"] += 1
                            elif health == 'OK':
                                cpu_count["✅"] += 1
                            else:
                                cpu_count["⚠️"] += 1
                    
                    # 상태 텍스트 업데이트
                    status_text = "CPU: "
                    if cpu_count["✅"] > 0:
                        status_text += f"{cpu_count['✅']}✅"
                    if cpu_count["⚠️"] > 0:
                        status_text += f"+{cpu_count['⚠️']}"
                    if cpu_count["❌"] > 0:
                        status_text += f" (❌{cpu_count['❌']})"
                    
                    # 상세보기 아이콘 추가
                    if cpu_count["❌"] > 0 or cpu_count["⚠️"] > 0:
                        status_text += " (상세보기 ℹ️)"
                    
                    self.status_labels['CPU'].setText(status_text)
        except Exception as e:
            self.status_labels['CPU'].setText("CPU: 오류")
            logger.error(f"CPU 상태 업데이트 실패: {e}")

    def _update_memory_status(self):
        try:
            if self.server_manager is not None:
                memory_data = self.server_manager.fetch_memory_info()
                if memory_data.get('Members@odata.count', 0) > 0:
                    mem_count = {"✅": 0, "❌": 0, "⚠️": 0}
                    total_capacity_gb = 0
                    
                    for member in memory_data.get('Members', []):
                        member_uri = member.get('@odata.id')
                        if member_uri:
                            memory_response = requests.get(
                                f"{self.server_manager.endpoints.base_url}{member_uri}",
                                auth=self.server_manager.auth,
                                verify=False
                            )
                            memory_info = memory_response.json()
                            status = memory_info.get('Status', {})
                            health = status.get('Health')
                            enabled = memory_info.get('Enabled', True)
                            
                            # 메모리 용량 계산 (MB를 GB로 변환)
                            capacity_mb = memory_info.get('CapacityMiB', 0)
                            if capacity_mb > 0:
                                total_capacity_gb += capacity_mb / 1024
                            
                            if not enabled or status.get('State') == 'Offline':
                                mem_count["❌"] += 1
                            elif health == 'OK':
                                mem_count["✅"] += 1
                            else:
                                mem_count["⚠️"] += 1
                    
                    # 상태 텍스트 업데이트 (총 용량 포함)
                    status_text = "MEM: "
                    if mem_count["✅"] > 0:
                        status_text += f"{mem_count['✅']}✅"
                    if mem_count["⚠️"] > 0:
                        status_text += f"+{mem_count['⚠️']}"
                    if mem_count["❌"] > 0:
                        status_text += f" (❌{mem_count['❌']})"
                    
                    # 총 용량 추가 (소수점 1자리까지)
                    if total_capacity_gb > 0:
                        status_text += f" ({total_capacity_gb:.1f}GB)"
                    
                    # 상세보기 아이콘 추가
                    if mem_count["❌"] > 0 or mem_count["⚠️"] > 0:
                        status_text += " (상세보기 ℹ️)"
                    
                    self.status_labels['MEM'].setText(status_text)
        except Exception as e:
            self.status_labels['MEM'].setText("MEM: 오류")
            logger.error(f"메모리 상태 업데이트 실패: {e}")

    def _update_disk_status(self):
        try:
            storage_data = self.server_manager.fetch_storage_info()
            if storage_data and 'Controllers' in storage_data:
                disk_count = {"✅": 0, "❌": 0, "⚠️": 0}
                
                for controller in storage_data.get('Controllers', []):
                    for drive in controller.get('Drives', []):
                        raid_status = drive.get('Oem', {}).get('Dell', {}).get('DellPhysicalDisk', {}).get('RaidStatus')
                        if raid_status == 'Online':
                            disk_count["✅"] += 1
                        elif raid_status == 'Failed':
                            disk_count["❌"] += 1
                        else:
                            disk_count["⚠️"] += 1
                
                status_parts = []
                for icon, count in disk_count.items():
                    if count > 0:
                        status_parts.append(f"{count}{icon}")
                
                status_text = "DSK: " + " ".join(status_parts)
                if disk_count["❌"] > 0 or disk_count["⚠️"] > 0:
                    status_text += " (상세보기 ℹ️)"
                self.status_labels['DSK'].setText(status_text)
            else:
                self.status_labels['DSK'].setText("DSK: --")
        except Exception as e:
            self.status_labels['DSK'].setText("DSK: 오류")
            logger.error(f"디스크 상태 업데이트 실패: {e}")

    def _update_psu_status(self):
        try:
            if self.server_manager is not None:
                power_data = self.server_manager.fetch_psu_info()
                if power_data and 'PowerSupplies' in power_data:
                    psu_count = {"✅": 0, "❌": 0, "⚠️": 0}
                    
                    for psu in power_data['PowerSupplies']:
                        status = psu.get('Status', {})
                        health = status.get('Health')
                        state = status.get('State')
                        
                        if state == 'Absent':
                            continue
                        elif state != 'Enabled' or health == 'Critical':
                            psu_count["❌"] += 1
                        elif health == 'OK':
                            psu_count["✅"] += 1
                        else:
                            psu_count["⚠️"] += 1
                    
                    # 상태 텍스트 업데이트
                    status_text = "PWR: "
                    if psu_count["✅"] > 0:
                        status_text += f"{psu_count['✅']}✅"
                    if psu_count["⚠️"] > 0:
                        status_text += f"+{psu_count['⚠️']}"
                    if psu_count["❌"] > 0:
                        status_text += f" (❌{psu_count['❌']})"
                    
                    # 상세보기 아이콘 추가
                    if psu_count["❌"] > 0 or psu_count["⚠️"] > 0:
                        status_text += " (상세보기 ℹ️)"
                    
                    self.status_labels['PWR'].setText(status_text)
        except Exception as e:
            self.status_labels['PWR'].setText("PWR: 오류")
            logger.error(f"전원 공급 장치 상태 업데이트 실패: {e}")

    def _update_status_label(self, component, statuses):
        if statuses:
            unique_statuses = set(status for _, status in statuses)
            if len(unique_statuses) == 1:
                status_text = f"{component}: {unique_statuses.pop()}"
            else:
                problem_items = [f"{item}: {status}" for item, status in statuses if status != 'OK' and status != 'None']
                status_text = f"{component}: " + (', '.join(problem_items) if problem_items else "OK")
        else:
            status_text = f"{component}: None"
        self.status_labels[component].setText(status_text)
        logger.debug(f"{component} 상태 업데이트: {status_text}")

    def show_component_details(self, component_type):
        try:
            if not self.server_manager:
                return
                
            info = {}
            if component_type == 'CPU':
                processors_data = self.server_manager.fetch_processors_info()
                if processors_data.get('Members@odata.count', 0) > 0:
                    for member in processors_data.get('Members', []):
                        member_uri = member.get('@odata.id')
                        if member_uri:
                            cpu_response = requests.get(
                                f"{self.server_manager.endpoints.base_url}{member_uri}",
                                auth=self.server_manager.auth,
                                verify=False
                            )
                            cpu_info = cpu_response.json()
                            cpu_num = cpu_info.get('Id', '').split('.')[-1]
                            info[f'CPU {cpu_num}'] = {
                                '모델': cpu_info.get('Model', 'N/A'),
                                '제조사': cpu_info.get('Manufacturer', 'N/A'),
                                '상태': '활성화' if cpu_info.get('Enabled', True) else '비활성화',
                                '코어 수': str(cpu_info.get('TotalCores', 'N/A')),
                                '스레드 수': str(cpu_info.get('TotalThreads', 'N/A')),
                                '최대 속도': f"{cpu_info.get('MaxSpeedMHz', 'N/A')}MHz",
                                '현재 속도': f"{cpu_info.get('OperatingSpeedMHz', 'N/A')}MHz"
                            }
                            
            elif component_type == 'MEM':
                memory_data = self.server_manager.fetch_memory_info()
                if memory_data.get('Members@odata.count', 0) > 0:
                    for member in memory_data.get('Members', []):
                        member_uri = member.get('@odata.id')
                        if member_uri:
                            memory_response = requests.get(
                                f"{self.server_manager.endpoints.base_url}{member_uri}",
                                auth=self.server_manager.auth,
                                verify=False
                            )
                            memory_info = memory_response.json()
                            dimm_id = memory_info.get('Id', '').split('/')[-1]
                            info[dimm_id] = {
                                '용량': f"{memory_info.get('CapacityMiB', 0) // 1024}GB",
                                '타입': memory_info.get('MemoryDeviceType', 'N/A'),
                                '속도': f"{memory_info.get('OperatingSpeedMhz', 'N/A')}MHz",
                                '상태': memory_info.get('Status', {}).get('Health', 'N/A'),
                                '제조사': memory_info.get('Manufacturer', 'N/A')
                            }

            elif component_type == 'DSK':
                storage_data = self.server_manager.fetch_storage_info()
                if storage_data and 'Controllers' in storage_data:
                    for controller in storage_data.get('Controllers', []):
                        for drive in controller.get('Drives', []):
                            drive_id = drive.get('Id', 'N/A')
                            dell_disk = drive.get('Oem', {}).get('Dell', {}).get('DellPhysicalDisk', {})
                            
                            info[drive_id] = {
                                '모델': drive.get('Model', 'N/A'),
                                '제조사': drive.get('Manufacturer', 'N/A'),
                                '용량': f"{drive.get('CapacityBytes', 0) // (1024**3)}GB",
                                '상태': dell_disk.get('RaidStatus', 'N/A'),
                                '미디어 타입': dell_disk.get('MediaType', 'N/A'),
                                '프로토콜': drive.get('Protocol', 'N/A')
                            }

            elif component_type == 'PWR':
                power_data = self.server_manager.fetch_psu_info()
                if power_data and 'PowerSupplies' in power_data:
                    for idx, psu in enumerate(power_data.get('PowerSupplies', [])):
                        psu_id = f'PSU {idx}'
                        info[psu_id] = {
                            '모델': psu.get('Model', 'N/A'),
                            '제조사': psu.get('Manufacturer', 'N/A'),
                            '파워': f"{psu.get('PowerCapacityWatts', 'N/A')}W",
                            '상태': psu.get('Status', {}).get('Health', 'N/A'),
                            '입력 전압': f"{psu.get('LastPowerOutputWatts', 'N/A')}V",
                            '효율성': psu.get('EfficiencyPercent', 'N/A')
                        }

            if info:
                dialog = DetailDialog(f"{component_type} 상세 정보", info, self)
                dialog.exec()
                
        except Exception as e:
            logger.error(f"{component_type} 상세 정보 표시 실패: {str(e)}")
            error_dialog = ErrorDialog(
                "상세 정보 오류",
                f"{component_type} 상세 정보를 가져오는 중 오류가 발생했습니다.",
                str(e),
                self
            )
            error_dialog.exec()

    def _on_quick_connect(self):
        """빠른 연결 버튼 클릭 핸들러"""
        from config.system.log_config import set_current_server
        logger.debug("빠른 연결 버튼 클릭")
        quick_connect_server = server_config.get_quick_connect_server()
        if quick_connect_server:
            # 로그에 현재 서버 설정
            set_current_server(quick_connect_server.NAME)
            try:
                self.quick_connect()
            finally:
                # 연결 시도 후 항상 시스템 상태로 복원 (성공/실패 여부와 무관)
                set_current_server('SYSTEM')

    def quick_connect(self):
        """빠른 연결 기능 - 설정된 빠른 연결 서버에 즉시 연결"""
        quick_connect_server = server_config.get_quick_connect_server()
        if quick_connect_server:
            logger.debug("빠른 연결 시도 시작")
            try:
                # 커서를 대기 상태로 변경
                QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
                
                # 진행 상태 다이얼로그
                # progress = QProgressDialog("빠른 연결 시도 중...", None, 0, 0, self)
                # progress.setWindowTitle("연결 중")
                # progress.setWindowModality(Qt.WindowModality.WindowModal)
                # progress.setCancelButton(None)
                # progress.show()
                
                # 메인 윈도우의 server_section을 통해 연결
                main_window = self.window()
                if hasattr(main_window, 'server_section'):
                    server_info = {
                        'NAME': quick_connect_server.NAME,
                        'IP': quick_connect_server.IP,
                        'PORT': quick_connect_server.PORT,
                        'USERNAME': quick_connect_server.USERNAME,
                        'PASSWORD': quick_connect_server.PASSWORD
                    }
                    main_window.server_section.connect_server(server_info)
                else:
                    logger.error("server_section을 찾을 수 없습니다.")
                    raise RuntimeError("서버 연결 기능을 찾을 수 없습니다.")
                    
            except Exception as e:
                logger.error(f"서버 연결 실패: {str(e)}")
                error_dialog = ErrorDialog(
                    "연결 오류",
                    "서버 연결 중 오류가 발생했습니다.",
                    str(e),
                    self
                )
                error_dialog.exec()
            finally:
                # 프로그레스 다이얼로그와 커서 정리 제거
                # progress.close() if 'progress' in locals() else None
                QApplication.restoreOverrideCursor()
        else:
            logger.warning("빠른 연결 서버가 설정되지 않았습니다.")
            error_dialog = ErrorDialog(
                "연결 오류",
                "빠른 연결 서버가 설정되지 않았습니다.",
                "서버 설정에서 빠른 연결 서버를 설정해주세요.",
                self
            )
            error_dialog.exec()


def get_all_system_settings(parent, server_manager):
    try:
        # 기본 정보 가져오기
        bios_info = server_manager.fetch_bios_info()
        idrac_info = server_manager.fetch_idrac_info()
        idrac_pwr_info = server_manager.fetch_idrac_pwr_info()
        nic_data = server_manager.fetch_network_adapters_info()
        
        # CPU 종류 확인 (bios_info에서 가져오기)
        attributes = bios_info.get('Attributes', {})
        cpu_brand = attributes.get('Proc1Brand', '')
        is_amd_cpu = 'AMD' in cpu_brand
        
        all_settings = {}
        
        # System Information
        all_settings['System Information'] = {
            'System Model Name': {'attr_name': 'SystemModelName', 'value': attributes.get('SystemModelName', 'N/A')},
            '제조사': {'attr_name': 'SystemManufacturer', 'value': attributes.get('SystemManufacturer', 'N/A')},
            'BIOS 버전': {'attr_name': 'SystemBiosVersion', 'value': attributes.get('SystemBiosVersion', 'N/A')},
            'System Service Tag': {'attr_name': 'SystemServiceTag', 'value': attributes.get('SystemServiceTag', 'N/A')}
        }
        
        # Processor Settings - CPU 종류에 따라 다르게 설정
        processor_settings = {
            'Logical Processor': {'attr_name': 'LogicalProc', 'value': attributes.get('LogicalProc', 'N/A')},
            'Virtualization Technology': {'attr_name': 'ProcVirtualization', 'value': attributes.get('ProcVirtualization', 'N/A')},
            'x2APIC Mode': {'attr_name': 'ProcX2Apic', 'value': attributes.get('ProcX2Apic', 'N/A')}
        }

        if is_amd_cpu:
            # AMD CPU 전용 설정
            processor_settings.update({
                'NUMA Nodes Per Socket': {'attr_name': 'NumaNodesPerSocket', 'value': attributes.get('NumaNodesPerSocket', 'N/A')},
                'L3 Cache as NUMA Domain': {'attr_name': 'L3CacheAsNumaDomain', 'value': attributes.get('L3CacheAsNumaDomain', 'N/A')},
                'MADT Core Enumeration': {'attr_name': 'MadtCoreEnumeration', 'value': attributes.get('MadtCoreEnumeration', 'N/A')},
                'Configurable TDP': {'attr_name': 'ConfigTdp', 'value': attributes.get('ConfigTdp', 'N/A')},
                'L1 Stream HW Prefetcher': {'attr_name': 'L1StreamHwPrefetcher', 'value': attributes.get('L1StreamHwPrefetcher', 'N/A')},
                'L2 Stream HW Prefetcher': {'attr_name': 'L2StreamHwPrefetcher', 'value': attributes.get('L2StreamHwPrefetcher', 'N/A')},
                'L1 Stride Prefetcher': {'attr_name': 'L1StridePrefetcher', 'value': attributes.get('L1StridePrefetcher', 'N/A')},
                'L1 Region Prefetcher': {'attr_name': 'L1RegionPrefetcher', 'value': attributes.get('L1RegionPrefetcher', 'N/A')},
                'L2 Up Down Prefetcher': {'attr_name': 'L2UpDownPrefetcher', 'value': attributes.get('L2UpDownPrefetcher', 'N/A')},
                'Secure Memory Encryption': {'attr_name': 'SecureMemoryEncryption', 'value': attributes.get('SecureMemoryEncryption', 'N/A')},
                'Minimum SEV non-ES ASID': {'attr_name': 'MinSevNonEsAsid', 'value': attributes.get('MinSevNonEsAsid', 'N/A')},
                'Secure Nested Paging': {'attr_name': 'SecureNestedPaging', 'value': attributes.get('SecureNestedPaging', 'N/A')},
                'SNP Memory Coverage': {'attr_name': 'SnpMemoryCoverage', 'value': attributes.get('SnpMemoryCoverage', 'N/A')},
                'Transparent Secure Memory Encryption': {'attr_name': 'TransparentSecureMemoryEncryption', 'value': attributes.get('TransparentSecureMemoryEncryption', 'N/A')},
            })
        else:
            # Intel CPU 전용 설정
            processor_settings.update({
                'Sub NUMA Cluster': {'attr_name': 'SubNumaCluster', 'value': attributes.get('SubNumaCluster', 'N/A')},
            })

        # x2APIC Mode는 공통 설정
        processor_settings.update({
            'x2APIC Mode': {'attr_name': 'ProcX2Apic', 'value': attributes.get('ProcX2Apic', 'N/A')}
        })

        all_settings['Processor Settings'] = processor_settings
        
        # Boot Settings
        all_settings['Boot Settings'] = {
            'Boot mode': {'attr_name': 'BootMode', 'value': attributes.get('BootMode', 'N/A')}
        }
        
        # Network Settings
        all_settings['Network Settings'] = {
            'PXE Device1': {'attr_name': 'PxeDev1EnDis', 'value': attributes.get('PxeDev1EnDis', 'N/A')},
            'PXE Device1 NIC': {'attr_name': 'PxeDev1Interface', 'value': attributes.get('PxeDev1Interface', 'N/A')},
            'PXE Device2': {'attr_name': 'PxeDev2EnDis', 'value': attributes.get('PxeDev2EnDis', 'N/A')},
            'PXE Device2 NIC': {'attr_name': 'PxeDev2Interface', 'value': attributes.get('PxeDev2Interface', 'N/A')},
            'PXE Device3': {'attr_name': 'PxeDev3EnDis', 'value': attributes.get('PxeDev3EnDis', 'N/A')},
            'PXE Device3 NIC': {'attr_name': 'PxeDev3Interface', 'value': attributes.get('PxeDev3Interface', 'N/A')},
            'PXE Device4': {'attr_name': 'PxeDev4EnDis', 'value': attributes.get('PxeDev4EnDis', 'N/A')},
            'PXE Device4 NIC': {'attr_name': 'PxeDev4Interface', 'value': attributes.get('PxeDev4Interface', 'N/A')}
        }
        
        # Integrated Devices
        all_settings['Integrated Devices'] = {
            'SR-IOV Global Enable': {'attr_name': 'SriovGlobalEnable', 'value': attributes.get('SriovGlobalEnable', 'N/A')},
            'OS Watchdog Timer': {'attr_name': 'OsWatchdogTimer', 'value': attributes.get('OsWatchdogTimer', 'N/A')}
        }

        # System Profile Settings
        system_profile_settings = {
            'System Profile': {'attr_name': 'SysProfile', 'value': attributes.get('SysProfile', 'N/A')},
            'CPU Power Management': {'attr_name': 'ProcPwrPerf', 'value': attributes.get('ProcPwrPerf', 'N/A')},
            'Memory Frequency': {'attr_name': 'ProcCStates', 'value': attributes.get('ProcCStates', 'N/A')},
            'C1E': {'attr_name': 'ProcC1E', 'value': attributes.get('ProcC1E', 'N/A')},
            'Turbo Boost': {'attr_name': 'ProcTurboMode', 'value': attributes.get('ProcTurboMode', 'N/A')},
            'Energy Efficiency Policy': {'attr_name': 'EnergyPerformanceBias', 'value': attributes.get('EnergyPerformanceBias', 'N/A')},
            'Memory Patrol Scrub': {'attr_name': 'MemPatrolScrub', 'value': attributes.get('MemPatrolScrub', 'N/A')}
        }

        # AMD CPU 전용 설정
        if is_amd_cpu:
            amd_settings = {
                'Determinism Slider': {'attr_name': 'DeterminismSlider', 'value': attributes.get('DeterminismSlider', 'N/A')},
                'Power Profile Select': {'attr_name': 'PowerProfileSelect', 'value': attributes.get('PowerProfileSelect', 'N/A')},
                'PCIE Speed PMM Control': {'attr_name': 'PCIESpeedPMMControl', 'value': attributes.get('PCIESpeedPMMControl', 'N/A')},
                'EQ Bypass To Highest Rate': {'attr_name': 'EQBypassToHighestRate', 'value': attributes.get('EQBypassToHighestRate', 'N/A')},
                'DF PState Frequency Optimizer': {'attr_name': 'DFPstateFrequencyOptimizer', 'value': attributes.get('DFPstateFrequencyOptimizer', 'N/A')},
                'DF PState Latency Optimizer': {'attr_name': 'DFPstateLatencyOptimizer', 'value': attributes.get('DFPstateLatencyOptimizer', 'N/A')},
                'DF CState': {'attr_name': 'DfCState', 'value': attributes.get('DfCState', 'N/A')},
                'Host System Management Port': {'attr_name': 'HSMPSupport', 'value': attributes.get('HSMPSupport', 'N/A')},
                'Boost FMax': {'attr_name': 'BoostFMax', 'value': attributes.get('BoostFMax', 'N/A')},
                'Algorithm Performance Boost Disable': {'attr_name': 'ApbDis', 'value': attributes.get('ApbDis', 'N/A')}
            }
            system_profile_settings.update(amd_settings)

        # System Profile Settings에 업데이트된 설정 할당
        all_settings['System Profile Settings'] = system_profile_settings

        # Miscellaneous Settings
        all_settings['Miscellaneous Settings'] = {
            'F1/F2 Prompt On Error': {'attr_name': 'ErrPrompt', 'value': attributes.get('ErrPrompt', 'N/A')}
        }

        # iDRAC Settings
        all_settings['iDRAC Settings'] = {
            'Mac Address': {'attr_name': 'NIC.1.MACAddress', 'value': idrac_info['Attributes'].get('NIC.1.MACAddress', 'N/A')},
            'Enable IPv4': {'attr_name': 'IPv4.1.Enable', 'value': idrac_info['Attributes'].get('IPv4.1.Enable', 'N/A')},
            'Enable DHCP': {'attr_name': 'IPv4.1.DHCPEnable', 'value': idrac_info['Attributes'].get('IPv4.1.DHCPEnable', 'N/A')},
            'Static IP Address': {'attr_name': 'IPv4Static.1.Address', 'value': idrac_info['Attributes'].get('IPv4Static.1.Address', 'N/A')},
            'Static Gateway': {'attr_name': 'IPv4Static.1.Gateway', 'value': idrac_info['Attributes'].get('IPv4Static.1.Gateway', 'N/A')},
            'Static Subnet Mask': {'attr_name': 'IPv4Static.1.Netmask', 'value': idrac_info['Attributes'].get('IPv4Static.1.Netmask', 'N/A')},
            'Enable IPMI Over LAN': {'attr_name': 'IPMILan.1.Enable', 'value': idrac_info['Attributes'].get('IPMILan.1.Enable', 'N/A')},
            'Enable VLAN ID': {'attr_name': 'NIC.1.VLanEnable', 'value': idrac_info['Attributes'].get('NIC.1.VLanEnable', 'N/A')}
        }

        # Power Configuration
        all_settings['Power Configuration'] = {
            'Redundancy Policy': {'attr_name': 'ServerPwr.1.PSRedPolicy', 'value': idrac_pwr_info['Attributes'].get('ServerPwr.1.PSRedPolicy', 'N/A')},
            'Enable Hot Spare': {'attr_name': 'ServerPwr.1.PSRapidOn', 'value': idrac_pwr_info['Attributes'].get('ServerPwr.1.PSRapidOn', 'N/A')}
        }

        # NIC Configuration
        all_settings['NIC Configuration'] = {}
        if nic_data and 'NetworkAdapters' in nic_data:
            for adapter in nic_data['NetworkAdapters']:
                for func in adapter.get('NetworkDeviceFunctions', []):
                    if func_id := func.get('Id'):
                        virt_info = server_manager.fetch_network_virtualization_info(
                            adapter.get('Id'), func_id)
                        if virt_info and 'Attributes' in virt_info:
                            attrs = virt_info['Attributes']
                            
                            # NIC 포트 ID 추가
                            all_settings['NIC Configuration'][func_id] = {
                                'attr_name': '',
                                'value': ''
                            }
                            # 각 설정을 순서대로 추가
                            all_settings['NIC Configuration'][f"{func_id}.VirtualizationMode"] = {
                                'attr_name': 'VirtualizationMode',
                                'value': attrs.get('VirtualizationMode', 'N/A')
                            }
                            all_settings['NIC Configuration'][f"{func_id}.LnkSpeed"] = {
                                'attr_name': 'LnkSpeed',
                                'value': attrs.get('LnkSpeed', 'N/A')
                            }
                            all_settings['NIC Configuration'][f"{func_id}.LegacyBootProto"] = {
                                'attr_name': 'LegacyBootProto',
                                'value': attrs.get('LegacyBootProto', 'N/A')
                            }

        return all_settings
    except Exception as e:
        if parent:
            QMessageBox.critical(parent, "오류", f"시스템 설정을 가져오는 중 오류가 발생했습니다: {str(e)}")
        return {}

def format_firmware_date(date_str):
    if not date_str or date_str == 'N/A':
        return 'N/A'
    try:
        # 예: "2024-03-22T20:33:49Z" -> "2024-03-22 20:33"
        if 'T' in date_str:
            date_parts = date_str.split('T')
            if len(date_parts) == 2:
                date = date_parts[0]  # "2024-03-22"
                time = date_parts[1][:5]  # "20:33"
                return f"{date} {time}"
        return date_str
    except Exception as e:
        logger.error(f"날짜 변환 중 오류 발생: {e}, date_str: {date_str}")
        return date_str

def save_system_info(parent_dialog, server_manager):
    if server_manager is None:
        error_dialog = ErrorDialog(
            "연결 오류",
            "서버에 연결되어 있는지 확인해주세요.",
            "서버 연결을 확인하고 다시 시도해주세요.",
            parent_dialog
        )
        error_dialog.exec()
        return

    # 진행률 다이얼로그 생성
    progress_dialog = QProgressDialog("시스템 정보 수집 중...", None, 0, 100, parent_dialog)
    progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
    progress_dialog.setWindowTitle("데이터 수집")
    progress_dialog.setAutoClose(True)
    progress_dialog.setMinimumDuration(0)
    progress_dialog.show()
    
    try:
        progress_dialog.setValue(0)
        progress_dialog.setLabelText("기본 시스템 정보 수집 중...")
        basic_info = server_manager.fetch_basic_info()
        service_tag = basic_info['system'].get('ServiceTag', 'Unknown')
        current_date = datetime.now().strftime('%Y%m%d')
        default_filename = f"{service_tag}_{current_date}.xlsx"
        
        documents_path = os.path.join(os.path.expanduser("~"), "Documents")
        default_path = os.path.join(documents_path, default_filename)

        progress_dialog.setValue(10)
        file_name, _ = QFileDialog.getSaveFileName(
            parent_dialog, 
            "시스템 설정 정보 저장",
            default_path,
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if file_name:
            if not file_name.endswith('.xlsx'):
                file_name += '.xlsx'

            model_name = basic_info['system'].get('Model', 'N/A')
            service_tag = basic_info['system'].get('ServiceTag', 'N/A')

            # 상태 정보 데이터 준비
            status_data = []
            
            progress_dialog.setValue(20)
            progress_dialog.setLabelText("CPU 정보 수집 중...")
            # CPU 정보 수집 (CPU Socket)
            processors_data = server_manager.fetch_processors_info()
            if processors_data.get('Members@odata.count', 0) > 0:
                status_data.append({
                    '구성 요소': 'CPU 정보',
                    'Dell Attribute name': '',
                    'value': ''
                })
                total_processors = len(processors_data.get('Members', []))
                for idx, member in enumerate(processors_data.get('Members', [])):
                    progress_value = 20 + (idx * 15 // total_processors)
                    progress_dialog.setValue(progress_value)
                    
                    member_uri = member.get('@odata.id')
                    if member_uri:
                        cpu_info = requests.get(
                            f"{server_manager.endpoints.base_url}{member_uri}",
                            auth=server_manager.auth,
                            verify=False
                        ).json()
                        
                        socket_id = cpu_info.get('Id', '').split('.')[-1]
                        status_data.extend([
                            {'구성 요소': f'    CPU CPU.Socket.{socket_id}', 'Dell Attribute name': '', 'value': ''},
                            {'구성 요소': '        모델', 'Dell Attribute name': 'Model', 'value': cpu_info.get('Model', 'N/A')},
                            {'구성 요소': '        제조사', 'Dell Attribute name': 'Manufacturer', 'value': cpu_info.get('Manufacturer', 'N/A')},
                            {'구성 요소': '        코어 수', 'Dell Attribute name': 'TotalCores', 'value': str(cpu_info.get('TotalCores', 'N/A'))},
                            {'구성 요소': '        스레드 수', 'Dell Attribute name': 'TotalThreads', 'value': str(cpu_info.get('TotalThreads', 'N/A'))},
                            {'구성 요소': '        최대 속도', 'Dell Attribute name': 'MaxSpeedMHz', 'value': str(cpu_info.get('MaxSpeedMHz', 'N/A'))},
                            {'구성 요소': '        현재 속도', 'Dell Attribute name': 'OperatingSpeedMHz', 'value': str(cpu_info.get('OperatingSpeedMHz', 'N/A'))},
                            {'구성 요소': '        상태', 'Dell Attribute name': 'Status.Health', 'value': cpu_info.get('Status', {}).get('Health', 'N/A')}
                        ])

            progress_dialog.setValue(35)
            progress_dialog.setLabelText("메모리 정보 수집 중...")
            # 메모리 정보 수집 (메모리 DIMM)
            memory_data = server_manager.fetch_memory_info()
            if memory_data.get('Members@odata.count', 0) > 0:
                status_data.append({
                    '구성 요소': '메모리 정보',
                    'Dell Attribute name': '',
                    'value': ''
                })
                
                memory_info_list = []
                total_memory = len(memory_data.get('Members', []))
                for idx, member in enumerate(memory_data.get('Members', [])):
                    progress_value = 35 + (idx * 15 // total_memory)
                    progress_dialog.setValue(progress_value)
                    
                    member_uri = member.get('@odata.id')
                    if member_uri:
                        memory_info = requests.get(
                            f"{server_manager.endpoints.base_url}{member_uri}",
                            auth=server_manager.auth,
                            verify=False
                        ).json()
                        memory_info_list.append(memory_info)
                
                memory_info_list.sort(key=lambda x: x.get('DeviceLocator', ''))
                
                for memory_info in memory_info_list:
                    dimm_socket = memory_info.get('DeviceLocator', 'Unknown')
                    if memory_info.get('CapacityMiB'):
                        status_data.extend([
                            {'구성 요소': f'    메모리 DIMM {dimm_socket}', 'Dell Attribute name': '', 'value': ''},
                            {'구성 요소': '        제조사', 'Dell Attribute name': 'Manufacturer', 'value': memory_info.get('Manufacturer', 'N/A')},
                            {'구성 요소': '        타입', 'Dell Attribute name': 'MemoryDeviceType', 'value': memory_info.get('MemoryDeviceType', 'N/A')},
                            {'구성 요소': '        동작 속도', 'Dell Attribute name': 'OperatingSpeedMhz', 'value': f"{memory_info.get('OperatingSpeedMhz', 'N/A')} MHz"},
                            {'구성 요소': '        용량', 'Dell Attribute name': 'CapacityMiB', 'value': convert_capacity(memory_info.get('CapacityMiB', 0), False)},
                            {'구성 요소': '        상태', 'Dell Attribute name': 'Status.Health', 'value': memory_info.get('Status', {}).get('Health', 'N/A')}
                        ])

            progress_dialog.setValue(50)
            progress_dialog.setLabelText("스토리지 정보 수집 중...")
            # 스토리지 정보 수집 (스토리지 컨트롤러, 볼륨)
            storage_data = server_manager.fetch_storage_info()
            if storage_data and 'Controllers' in storage_data:
                status_data.append({
                    '구성 요소': '스토리지 정보',
                    'Dell Attribute name': '',
                    'value': ''
                })

                # RAID 컨트롤러 정보
                for controller in storage_data.get('Controllers', []):
                    controller_id = controller.get('Id', '')
                    storage_controllers = controller.get('StorageControllers', [])
                    if storage_controllers:
                        controller_info = storage_controllers[0]
                        
                        cache_size = controller_info.get('CacheSummary', {}).get('TotalCacheSizeMiB', 0)
                        cache_value = convert_capacity(cache_size, False)
                        
                        status_data.extend([
                            {'구성 요소': f'컨트롤러 {controller_id}', 'Dell Attribute name': '', 'value': ''},
                            {'구성 요소': '    모델', 'Dell Attribute name': 'Model', 'value': controller_info.get('Model', 'N/A')},
                            {'구성 요소': '    펌웨어 버전', 'Dell Attribute name': 'FirmwareVersion', 'value': controller_info.get('FirmwareVersion', 'N/A')},
                            {'구성 요소': '    캐시', 'Dell Attribute name': 'TotalCacheSizeMiB', 'value': cache_value},
                            {'구성 요소': '    상태', 'Dell Attribute name': 'Status.Health', 'value': controller.get('Status', {}).get('Health', 'N/A')}
                        ])

                        # 볼륨 정보
                        volumes = controller.get('Volumes', [])
                        if volumes:
                            for volume in volumes:
                                volume_name = volume.get('Name', '')  # Disk.Virtual.239:RAID.SL.3-1 형식
                                dell_volume = volume.get('Oem', {}).get('Dell', {}).get('DellVolume', {})
                                
                                status_data.extend([
                                    {'구성 요소': f'    볼륨: {volume_name}', 'Dell Attribute name': '', 'value': ''},
                                    {'구성 요소': '        레이드', 'Dell Attribute name': 'RAIDType', 'value': volume.get('RAIDType', 'N/A')},
                                    {'구성 요소': '        미디어 타입', 'Dell Attribute name': 'MediaType', 'value': dell_volume.get('MediaType', 'N/A')},
                                    {'구성 요소': '        용량', 'Dell Attribute name': 'CapacityBytes', 'value': convert_capacity(volume.get('CapacityBytes', 0), True)},
                                    {'구성 요소': '        상태', 'Dell Attribute name': 'RaidStatus', 'value': dell_volume.get('RaidStatus', 'N/A')}
                                ])

                                # 드라이브 정보 수집 및 정렬
                                drives_info = []
                                volume_drives = volume.get('Links', {}).get('Drives', [])

                                for drive_link in volume_drives:
                                    drive_uri = drive_link.get('@odata.id')
                                    if drive_uri:
                                        drive_info = requests.get(
                                            f"{server_manager.endpoints.base_url}{drive_uri}",
                                            auth=server_manager.auth,
                                            verify=False
                                        ).json()
                                        drives_info.append(drive_info)

                                # Bay 번호로 정렬 (Bay 형식이 아닌 경우 원래 순서 유지)
                                def get_bay_number(drive_info):
                                    drive_id = drive_info.get('Id', '')
                                    try:
                                        if 'Bay.' in drive_id:
                                            return int(drive_id.split('Bay.')[-1].split(':')[0])
                                        return 0
                                    except (ValueError, IndexError):
                                        return 0

                                drives_info.sort(key=get_bay_number)

                                # 정렬된 드라이브 정보 추가
                                for drive in drives_info:
                                    drive_id = drive.get('Id', '')
                                    # 간단한 형식으로 드라이브 ID 변환 (Disk.Bay.X 형식만 추출)
                                    simplified_id = drive_id.split(':')[0]
                                    dell_drive = drive.get('Oem', {}).get('Dell', {}).get('DellPhysicalDisk', {})
                                    
                                    status_data.extend([
                                        {'구성 요소': f'        드라이브: {simplified_id}', 'Dell Attribute name': '', 'value': ''},
                                        {'구성 요소': '            제조사', 'Dell Attribute name': 'Manufacturer', 'value': drive.get('Manufacturer', 'N/A')},
                                        {'구성 요소': '            파트 번호', 'Dell Attribute name': 'PartNumber', 'value': drive.get('PartNumber', 'N/A')},
                                        {'구성 요소': '            시리얼 번호', 'Dell Attribute name': 'SerialNumber', 'value': drive.get('SerialNumber', 'N/A')},
                                        {'구성 요소': '            용량', 'Dell Attribute name': 'CapacityBytes', 'value': convert_capacity(drive.get('CapacityBytes', 0), True)},
                                        {'구성 요소': '            레이드 상태', 'Dell Attribute name': 'RaidStatus', 'value': dell_drive.get('RaidStatus', 'N/A')}
                                    ])

            progress_dialog.setValue(60)
            progress_dialog.setLabelText("NIC 정보 수집 중...")
            # NIC 정보 수집 (네트워크 어댑터)
            nic_data = server_manager.fetch_network_adapters_info()
            if nic_data and 'NetworkAdapters' in nic_data:
                status_data.append({
                    '구성 요소': 'NIC 정보',
                    'Dell Attribute name': '',
                    'value': ''
                })
                
                # NIC 타입별 우선순위 설정
                nic_type_order = {'NIC.Embedded.': 0, 'NIC.Integrated.': 1, 'NIC.Slot.': 2}
                
                # NIC 정보 정렬
                sorted_adapters = sorted(
                    nic_data['NetworkAdapters'],
                    key=lambda x: (
                        next((order for type_key, order in nic_type_order.items() if type_key in x.get('Id', '')), 3),
                        *(map(lambda n: int(n) if n.isdigit() else 0, x.get('Id', '').split('.')[-1].split('-')))
                    )
                )
                
                # iDRAC 라이센스 확인
                license_info = server_manager.check_idrac_license()
                
                for adapter in sorted_adapters:
                    adapter_id = adapter.get('Id', 'Unknown')
                    status_data.extend([
                        {'구성 요소': f'    NIC 어댑터: {adapter_id}', 'Dell Attribute name': '', 'value': ''},
                        {'구성 요소': '        모델', 'Dell Attribute name': 'Model', 'value': adapter.get('Model', 'N/A')},
                        {'구성 요소': '        제조사', 'Dell Attribute name': 'Manufacturer', 'value': adapter.get('Manufacturer', 'N/A')},
                        {'구성 요소': '        파트 번호', 'Dell Attribute name': 'PartNumber', 'value': adapter.get('PartNumber', 'N/A')},
                        {'구성 요소': '        시리얼 번호', 'Dell Attribute name': 'SerialNumber', 'value': adapter.get('SerialNumber', 'N/A')},
                        {'구성 요소': '        상태', 'Dell Attribute name': 'Status.Health', 'value': adapter.get('Status', {}).get('Health', 'N/A')}
                    ])

                    # 컨트롤러 정보
                    for controller in adapter.get('Controllers', []):
                        controller_item = ['        컨트롤러 정보', '', '']
                        status_data.extend([
                            {'구성 요소': '        컨트롤러 정보', 'Dell Attribute name': '', 'value': ''},
                            {'구성 요소': '            펌웨어 버전', 'Dell Attribute name': 'FirmwarePackageVersion', 'value': controller.get('FirmwarePackageVersion', 'N/A')},
                            {'구성 요소': '            가상화 지원', 'Dell Attribute name': 'VirtualizationOffload', 
                            'value': '가상화 미지원 카드' if not controller.get('ControllerCapabilities', {}).get('VirtualizationOffload', {}).get('SRIOV', {}).get('SRIOVVEPACapable') 
                            else '가상화 지원 카드'}
                        ])

                        # 포트 정보
                        for port in adapter.get('NetworkPorts', []):
                            port_id = port.get('Id', 'N/A')
                            
                            port_info = port
                            device_function_id = f"{port_id}-1"
                            
                            device_function_info = requests.get(
                                f"{server_manager.endpoints.base_url}/redfish/v1/Chassis/System.Embedded.1/NetworkAdapters/{adapter.get('Id')}/NetworkDeviceFunctions/{device_function_id}/Oem/Dell/DellNetworkAttributes/{device_function_id}",
                                auth=server_manager.auth,
                                verify=False
                            ).json()
                            
                            virtualization_mode = device_function_info.get('Attributes', {}).get('VirtualizationMode', 'N/A')

                            status_data.extend([
                                {'구성 요소': f'        포트: {port_id}', 'Dell Attribute name': '', 'value': ''},
                                {'구성 요소': '            링크 상태', 'Dell Attribute name': 'LinkStatus', 'value': port_info.get('LinkStatus', 'N/A')},
                                {'구성 요소': '            현재 속도', 'Dell Attribute name': 'CurrentLinkSpeedMbps', 'value': f"{port_info.get('CurrentLinkSpeedMbps', 'N/A')} Mbps"},
                                {'구성 요소': '            Flow Control 설정', 'Dell Attribute name': 'FlowControlConfiguration', 'value': port_info.get('FlowControlConfiguration', 'N/A')},
                                {'구성 요소': '            Flow Control 상태', 'Dell Attribute name': 'FlowControlStatus', 'value': port_info.get('FlowControlStatus', 'N/A')},
                                {'구성 요소': '            MAC 주소', 'Dell Attribute name': 'AssociatedNetworkAddresses', 'value': port_info.get('AssociatedNetworkAddresses', ['N/A'])[0] if port_info.get('AssociatedNetworkAddresses') else 'N/A'},
                                {'구성 요소': '            가상화 모드', 'Dell Attribute name': 'VirtualizationMode', 'value': virtualization_mode}
                            ])
                        # 트랜시버 정보 (Enterprise 라이센스 이상)
                        transceiver = port.get('Oem', {}).get('Dell', {}).get('DellNetworkTransceiver', {})
                        if transceiver and license_info and 'enterprise' in license_info['type'].lower():
                            status_data.extend([
                                {'구성 요소': '            트랜시버 정보', 'Dell Attribute name': '', 'value': ''},
                                {'구성 요소': '                트랜시버 타입', 'Dell Attribute name': 'IdentifierType', 'value': transceiver.get('IdentifierType', 'N/A')},
                                {'구성 요소': '                인터페이스', 'Dell Attribute name': 'InterfaceType', 'value': transceiver.get('InterfaceType', 'N/A')},
                                {'구성 요소': '                트랜시버 제조사', 'Dell Attribute name': 'VendorName', 'value': transceiver.get('VendorName', 'N/A')}
                            ])

                            # 광 레벨 정보 (Datacenter 라이센스)
                            if 'datacenter' in license_info['type'].lower():
                                optical_data = transceiver.get('OpticalData', {})
                                if optical_data:
                                    status_data.extend([
                                        {'구성 요소': '                광 레벨 정보', 'Dell Attribute name': '', 'value': ''},
                                        {'구성 요소': '                    온도', 'Dell Attribute name': 'Temperature', 'value': f"{optical_data.get('Temperature', 'N/A')} °C"},
                                        {'구성 요소': '                    전압', 'Dell Attribute name': 'SupplyVoltage', 'value': f"{optical_data.get('SupplyVoltage', 'N/A')} V"},
                                        {'구성 요소': '                    TX 파워', 'Dell Attribute name': 'TxPower', 'value': f"{optical_data.get('TxPower', 'N/A')} dBm"},
                                        {'구성 요소': '                    RX 파워', 'Dell Attribute name': 'RxPower', 'value': f"{optical_data.get('RxPower', 'N/A')} dBm"},
                                        {'구성 요소': '                    레이저 바이어스 전류', 'Dell Attribute name': 'LaserBiasCurrent', 'value': f"{optical_data.get('LaserBiasCurrent', 'N/A')} mA"}
                                    ])

            progress_dialog.setValue(70)
            progress_dialog.setLabelText("PSU 정보 수집 중...")
            # PSU 정보 수집 (전원)
            power_data = server_manager.fetch_psu_info()
            if power_data and 'PowerSupplies' in power_data:
                status_data.append({
                    '구성 요소': 'PSU 정보',
                    'Dell Attribute name': '',
                    'value': ''
                })
                for idx, psu in enumerate(power_data.get('PowerSupplies', [])):
                    psu_id = f'PSU {idx + 1}'
                    status_data.extend([
                        {'구성 요소': f'    {psu_id}', 'Dell Attribute name': '', 'value': ''},
                        {'구성 요소': f'        모델', 'Dell Attribute name': 'Model', 'value': psu.get('Model', 'N/A')},
                        {'구성 요소': f'        제조사', 'Dell Attribute name': 'Manufacturer', 'value': psu.get('Manufacturer', 'N/A')},
                        {'구성 요소': f'        파워', 'Dell Attribute name': 'PowerCapacityWatts', 'value': f"{psu.get('PowerCapacityWatts', 'N/A')}W"},
                        {'구성 요소': f'        상태', 'Dell Attribute name': 'Status.Health', 'value': psu.get('Status', {}).get('Health', 'N/A')},
                        {'구성 요소': f'        입력 전압', 'Dell Attribute name': 'LastPowerOutputWatts', 'value': f"{psu.get('LastPowerOutputWatts', 'N/A')}V"},
                    ])

            progress_dialog.setValue(80)
            progress_dialog.setLabelText("iDRAC MAC 주소 정보 수집 중...")
            # iDRAC MAC 주소 데이터 수집 (iDRAC MAC 주소)
            idrac_data = server_manager.fetch_detailed_info(server_manager.endpoints.idrac_mac_address)
            if idrac_data and 'Attributes' in idrac_data:
                status_data.append({
                    '구성 요소': 'iDRAC MAC 주소 정보',
                    'Dell Attribute name': '',
                    'value': ''
                })
                
                mac_address = idrac_data['Attributes'].get('CurrentNIC.1.MACAddress', 'N/A')
                status_data.append({
                    '구성 요소': '    MAC 주소',
                    'Dell Attribute name': 'CurrentNIC.1.MACAddress',
                    'value': mac_address
                })

            progress_dialog.setValue(90)
            progress_dialog.setLabelText("Excel 파일 생성 중...")
            # Excel 파일 생성 및 저장 로직
            bios_data = []
            
            for section_name, section_data in get_all_system_settings(parent_dialog, server_manager).items():
                bios_data.append({
                    'Settings': section_name,
                    'Dell Attribute name': '',
                    'value': ''
                })
                
                if isinstance(section_data, dict):
                    for setting_name, setting_info in section_data.items():
                        if isinstance(setting_info, dict):
                            bios_data.append({
                                'Settings': setting_name,
                                'Dell Attribute name': setting_info['attr_name'],
                                'value': setting_info['value']
                            })
                        else:
                            bios_data.append({
                                'Settings': '',
                                'Dell Attribute name': setting_name,
                                'value': setting_info
                            })
            
            firmware_data = server_manager.fetch_firmware_inventory()
            firmware_rows = []
            
            if firmware_data:
                # 기본 정보 가져오기
                basic_info = server_manager.fetch_basic_info()
                model_name = basic_info['system'].get('Model', 'N/A')
                service_tag = basic_info['system'].get('ServiceTag', 'N/A')
                
                # 모델명과 서비스태그만 첫 행에 추가
                firmware_rows.append({
                    '카테고리': model_name,
                    '장치명': service_tag,
                    '현재 버전': '',
                    '업데이트 날짜': ''
                })
                
                # 헤더 행 추가
                firmware_rows.append({
                    '카테고리': '카테고리',
                    '장치명': '장치명',
                    '현재 버전': '현재 버전',
                    '업데이트 날짜': '업데이트 날짜'
                })
                
                # 컴포넌트 정보 수집
                components = {
                    "BIOS 펌웨어": None,
                    "iDRAC 펌웨어": None,
                    "RAID": None,
                    "NIC": [],
                    "HBA": [],
                }
                
                # 펌웨어 그룹 초기화
                firmware_groups = {
                    'BIOS': {'installed': None, 'previous': None},
                    'iDRAC': {'installed': None, 'previous': None},
                    'RAID': {'installed': None, 'previous': None},
                    'NIC': [],
                    'HBA': []
                }

                # 펌웨어 데이터를 그룹별로 분류
                for member in firmware_data.get('Members', []):
                    if member_uri := member.get('@odata.id'):
                        component_id = member_uri.split('/')[-1]
                        component_info = server_manager.fetch_firmware_component(component_id)
                        
                        # Installed 버전과 Previous 버전 구분
                        version_type = 'installed' if component_id.startswith('Installed-') else 'previous'
                        
                        if 'BIOS' in component_id:
                            firmware_groups['BIOS'][version_type] = component_info
                        elif 'iDRAC' in component_id:
                            firmware_groups['iDRAC'][version_type] = component_info
                        elif 'PERC' in component_info.get('Name', ''):
                            firmware_groups['RAID'][version_type] = component_info
                        elif 'FC.' in component_id or 'QLogic' in component_info.get('Name', ''):
                            firmware_groups['HBA'].append(component_info)
                        elif 'NIC' in component_id:
                            firmware_groups['NIC'].append(component_info)

                # BIOS, iDRAC, RAID 정보 추가 (installed 버전 우선)
                for category, versions in [('BIOS 펌웨어', firmware_groups['BIOS']), 
                                        ('iDRAC 펌웨어', firmware_groups['iDRAC']), 
                                        ('RAID', firmware_groups['RAID'])]:
                    component = versions['installed'] or versions['previous']  # installed가 없으면 previous 사용
                    if component:
                        name = component.get('Name', 'N/A')
                        version = component.get('Version', 'N/A')
                        install_date = component.get('Oem', {}).get('Dell', {}).get(
                            'DellSoftwareInventory', {}).get('InstallationDate', 'N/A')
                        
                        firmware_rows.append({
                            '카테고리': category,
                            '장치명': name,
                            '현재 버전': version,
                            '업데이트 날짜': format_firmware_date(install_date)
                        })

                # HBA 정보 추가
                if firmware_groups["HBA"]:
                    firmware_rows.append({
                        '카테고리': 'HBA Card',
                        '장치명': '',
                        '현재 버전': '',
                        '업데이트 날짜': ''
                    })
                    
                    hba_versions = {}
                    for component in firmware_groups["HBA"]:
                        name = component.get('Name', 'N/A')
                        if ' - ' in name:
                            name = name.split(' - ')[0]
                        
                        version = component.get('Version', 'N/A')
                        install_date = component.get('Oem', {}).get('Dell', {}).get('DellSoftwareInventory', {}).get('InstallationDate', 'N/A')
                        
                        if name not in hba_versions or version > hba_versions[name]['version']:
                            hba_versions[name] = {
                                'version': version,
                                'date': install_date
                            }
                    
                    for name, info in hba_versions.items():
                        firmware_rows.append({
                            '카테고리': '',
                            '장치명': name,
                            '현재 버전': info['version'],
                            '업데이트 날짜': format_firmware_date(info['date'])
                        })

                # NIC 정보 추가
                if firmware_groups["NIC"]:
                    # NIC 버전 정보를 임시로 저장할 딕셔너리
                    nic_versions = {}
                    
                    # 각 NIC의 최신 버전 정보 수집
                    for component in firmware_groups["NIC"]:
                        name = component.get('Name', 'N/A')
                        if ' - ' in name:
                            name = name.split(' - ')[0]
                        
                        version = component.get('Version', 'N/A')
                        install_date = component.get('Oem', {}).get('Dell', {}).get('DellSoftwareInventory', {}).get('InstallationDate', 'N/A')
                        
                        # 기존 버전과 비교하여 최신 버전만 유지
                        if name not in nic_versions or version > nic_versions[name]['version']:
                            nic_versions[name] = {
                                'component': component,
                                'version': version,
                                'date': install_date
                            }
                    
                    # NIC 카테고리 행 추가
                    firmware_rows.append({
                        '카테고리': 'NIC',
                        '장치명': '',
                        '현재 버전': '',
                        '업데이트 날짜': ''
                    })
                    
                    # 최신 버전의 NIC 정보만 추가
                    for name, info in nic_versions.items():
                        firmware_rows.append({
                            '카테고리': '',  # NIC 아래 항목은 빈 카테고리
                            '장치명': name,
                            '현재 버전': info['version'],
                            '업데이트 날짜': format_firmware_date(info['date'])
                        })
            
                # 데이터프레임 생성 및 엑셀 파일 저장
                df = pd.DataFrame(firmware_rows)
                with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                    # 상태 정보 시트 작성
                    df_status = pd.DataFrame(status_data)
                    df_status.to_excel(writer, sheet_name='상태 정보', index=False)
                    
                    # 상태 정보 시트 스타일 적용
                    status_worksheet = writer.sheets['상태 정보']
                    
                    # 상태 정보 헤더 스타일 적용
                    header_font = Font(bold=True)
                    for cell in status_worksheet[1]:
                        cell.font = header_font
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                    
                    # 상태 정보 내용 스타일 적용
                    for row in status_worksheet.iter_rows(min_row=2):
                        # CPU 카테고리 및 하위 항목 구분을 위한 스타일
                        if row[0].value and not row[1].value and not row[2].value:
                            row[0].font = header_font
                            row[0].fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                        
                        # 모든 셀에 테두리 적용
                        for cell in row:
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                    
                    # 열 너비 자동 조정
                    for column in status_worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = max_length + 2
                        if any(ord(c) > 127 for c in str(column[0].value)):
                            adjusted_width = adjusted_width * 1.5
                        status_worksheet.column_dimensions[column_letter].width = adjusted_width

                    # BIOS 정보 시트 작성
                    df_bios = pd.DataFrame(bios_data)
                    df_bios.to_excel(writer, sheet_name='BIOS 정보', index=False)
                    
                    # BIOS 정보 시트 스타일 적용
                    bios_worksheet = writer.sheets['BIOS 정보']
                    
                    # BIOS 정보 헤더 굵게 설정
                    header_font = Font(bold=True)
                    for cell in bios_worksheet[1]:  # 첫 번째 행의 모든 셀
                        cell.font = header_font
                    
                    # Settings, Dell Attribute name, value가 비어있는 셀 굵게 설정
                    for row in bios_worksheet.iter_rows(min_row=2):
                        if not row[1].value or not row[2].value:
                            row[0].font = header_font
                    
                    # BIOS 정보 시트 열 너비 조정
                    for column in bios_worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = max_length + 2
                        if any(ord(c) > 127 for c in str(column[0].value)):
                            adjusted_width = adjusted_width * 1.5
                        bios_worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # BIOS 정보 시트 테두리 적용
                    for row in bios_worksheet.iter_rows(min_row=1, max_row=bios_worksheet.max_row):
                        for cell in row:
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                    
                    # 펌웨어 정보 시트 작성
                    df = pd.DataFrame(firmware_rows)
                    df.to_excel(writer, sheet_name='펌웨어 정보', index=False)
                    
                    # 펌웨어 정보 시트 가져오기
                    worksheet = writer.sheets['펌웨어 정보']
                    
                    # 펌웨어 정보 헤더 굵게 설정 (카테고리, 장치명, 현재 버전, 업데이트 날짜)
                    header_row = 2  # 두 번째 행이 실제 헤더
                    for cell in worksheet[header_row]:
                        cell.font = Font(bold=True)
                    
                    # 첫 번째 행 (모델명과 서비스태그)만 A1, B1에 배치
                    worksheet['A1'] = model_name
                    worksheet['B1'] = service_tag
                    worksheet['C1'] = ''
                    worksheet['D1'] = ''
                    
                    # 나머지 데이터는 두 번째 행부터 시작
                    for idx, row in enumerate(df.iloc[1:].itertuples(), start=2):
                        for col, value in enumerate(row[1:], start=1):
                            worksheet.cell(row=idx, column=col, value=value)
                    
                    # NIC 카테고리 병합
                    nic_start = None
                    for idx, row in enumerate(worksheet.iter_rows(min_row=3), start=3):
                        if row[0].value == 'NIC':
                            nic_start = idx
                            break
                    
                    if nic_start:
                        worksheet.merge_cells(f'A{nic_start}:A{worksheet.max_row}')
                    
                    # 펌웨어 정보 시트 스타일 적용
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 펌웨어 정보 시트 열 너비 조정
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = max_length + 2
                        if any(ord(c) > 127 for c in str(column[0].value)):
                            adjusted_width = adjusted_width * 1.5
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # 펌웨어 정보 시트 테두리 적용
                    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                        for cell in row:
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                    
                    # 헤더 행 스타일 적용
                    for cell in worksheet[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                    
                    # 섹션 헤더 스타일 적용 (카테고리만 굵게)
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                        if row[0].value and not row[1].value:  # 카테고리 값이 있고 장치명이 비어있는 경우
                            for cell in row:
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
            
            progress_dialog.setValue(100)
            progress_dialog.setLabelText("저장 완료")

            success_dialog = ErrorDialog(
                "저장 완료",
                "엑셀 파일이 저장되었습니다.",
                f"파일 위치: {os.path.abspath(file_name)}",
                parent_dialog
            )
            success_dialog.exec()
    except Exception as e:
        logger.error(f"시스템 설정 정보 저장 실패: {str(e)}")
        error_dialog = ErrorDialog(
            "시스템 설정 정보 저장 실패",
            "시스템 정보 저장 중 오류가 발생했습니다.",
            str(e),
            parent_dialog
        )
        error_dialog.exec()
    finally:
        progress_dialog.close()

def create_hardware_section(parent=None):
    return HardwareInfoWidget(parent)
