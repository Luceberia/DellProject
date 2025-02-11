import base64
import os
import time
from pathlib import Path
import requests
import openpyxl
from openpyxl.styles import Font
from collections import Counter
from datetime import datetime
import matplotlib
matplotlib.use('qtagg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from utils.system_utils import get_system_matplotlib_font

from config.system.log_config import setup_logging
from managers.dell_server_manager import DellServerManager
from PyQt6.QtCore import Qt, QTimer, QSettings
from PyQt6.QtGui import QColor, QIcon, QImage, QPixmap
from PyQt6.QtWidgets import (QApplication, QCheckBox, QComboBox, QDialog, QDialogButtonBox, 
                             QFileDialog, QGroupBox, QHBoxLayout, QLabel, QLineEdit, 
                             QMainWindow, QMenu, QMessageBox, QPushButton, QProgressBar, 
                             QProgressDialog, QSpinBox, QTreeWidget, QTreeWidgetItem, QVBoxLayout, QTabWidget, QFileDialog, QWidget, QScrollArea, QTableWidget, QTableWidgetItem)
from typing import Optional, cast
from ui.components.popups.error_dialog import ErrorDialog
from utils.utils import convert_capacity
from utils.cafe24 import cafe24_manager

logger = setup_logging()

def create_section(title, items, parent=None):
    group = QGroupBox(title)
    layout = QVBoxLayout(group)
    layout.setSpacing(10)
    layout.setContentsMargins(10, 15, 10, 15)
    
    # 섹션별 이모티콘 매핑
    icons = {
        # 모니터링 섹션
        "시스템 상태": "📊",
        "펌웨어 정보": "📦",
        # 관리 섹션
        "BIOS 설정": "🔧",
        "SSH 연결": "🔌",
        # 로그 섹션
        "LC LOG": "📜",
        "TSR LOG": "📋",
    }
    
    buttons = {}
    for item in items:
        # 해당 항목의 이모티콘 가져오기 (없으면 기본 화살표)
        icon = icons.get(item, "▸")
        btn = QPushButton(f"{icon} {item}")
        layout.addWidget(btn)
        buttons[item] = btn
        
        # 시스템 상태 버튼 클릭 이벤트 처리
        if item == "시스템 상태":
            btn.clicked.connect(lambda checked=False, p=parent: show_all_status(p))
        # 펌웨어 정보 버튼 클릭 이벤트 처리
        elif item == "펌웨어 정보":
            btn.clicked.connect(lambda checked=False, p=parent: show_firmware_info(p))
        # BIOS 설정 버튼 클릭 이벤트 처리
        elif item == "BIOS 설정":
            btn.clicked.connect(lambda checked=False, p=parent: show_system_info(p))
        # 작업 관리 버튼 클릭 이벤트 처리
        elif item == "작업 관리":
            btn.clicked.connect(lambda checked=False, p=parent: show_task_manager(p))
        # SSH 연결 버튼 클릭 이벤트 처리
        elif item == "SSH 연결":
            btn.clicked.connect(lambda checked=False, p=parent: open_ssh_connection(p))
        # LC LOG와 TSR LOG 버튼 클릭 이벤트 처리
        elif item == "LC LOG":
            btn.clicked.connect(lambda checked=False, p=parent: show_lc_log_popup(p))
        elif item == "TSR LOG":
            btn.clicked.connect(lambda checked=False, p=parent: show_tsr_log_popup(p))

    return group, buttons

def get_main_window() -> Optional[QMainWindow]:
    """메인 윈도우 객체 가져오기"""
    app = cast(QApplication, QApplication.instance())
    if app is not None:
        window = app.activeWindow()
        if isinstance(window, QMainWindow):
            return window
    return None

def create_monitor_section(parent=None):
    monitor_group = QGroupBox()
    monitor_layout = QHBoxLayout(monitor_group)
    monitor_layout.setSpacing(5)
    
    sections = {
        "📊 모니터링": ["시스템 상태", "펌웨어 정보"],
        "⚙️ 관리": ["BIOS 설정", "작업 관리", "SSH 연결"],
        "📋 로그": ["LC LOG", "TSR LOG"]
    }
    
    for title, items in sections.items():
        section_group, buttons = create_section(title, items, monitor_group)
        monitor_layout.addWidget(section_group)
    
    return monitor_group

def show_hostname_input_dialog(parent):
    dialog = QDialog(parent)
    dialog.setWindowTitle("호스트네임 변경")
    
    layout = QVBoxLayout()
    
    # 호스트네임 입력 필드
    hostname_label = QLabel("새로운 호스트네임:")
    hostname_input = QLineEdit()
    layout.addWidget(hostname_label)
    layout.addWidget(hostname_input)
    
    # 호스트네임 초기화 체크박스
    clear_hostname = QCheckBox("호스트네임 초기화 (빈 값으로 설정)")
    clear_hostname.toggled.connect(lambda checked: hostname_input.setEnabled(not checked))
    layout.addWidget(clear_hostname)
    
    # 확인/취소 버튼
    button_box = QDialogButtonBox(
        QDialogButtonBox.StandardButton.Ok | 
        QDialogButtonBox.StandardButton.Cancel
    )
    button_box.accepted.connect(dialog.accept)
    button_box.rejected.connect(dialog.reject)
    layout.addWidget(button_box)
    
    dialog.setLayout(layout)
    
    if dialog.exec() == QDialog.DialogCode.Accepted:
        if clear_hostname.isChecked():
            return ""  # 빈 문자열 반환
        return hostname_input.text()
    return None

def show_options(item):
    """옵션 UI를 표시합니다."""
    try:
        command_info = item.data(0, Qt.ItemDataRole.UserRole)
        if command_info and command_info.get('has_options'):
            # 옵션 위젯 찾기
            dialog = item.treeWidget().window()
            options_widget = dialog.findChild(QWidget, "options_widget")
            if not options_widget:
                return
                
            options_layout = options_widget.layout()
            
            # 기존 옵션들 제거
            while options_layout.count():
                child = options_layout.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()
            
            # 새 옵션들 추가
            for option in command_info['options']:
                if option.get('needs_input'):
                    # 입력이 필요한 옵션은 수평 레이아웃 사용
                    option_widget = QWidget()
                    option_layout = QHBoxLayout()
                    option_layout.setContentsMargins(0, 0, 0, 0)
                    
                    checkbox = QCheckBox(option['label'])
                    checkbox.setProperty('value', option['value'])
                    checkbox.setProperty('needs_input', True)
                    
                    input_field = QLineEdit()
                    input_field.setPlaceholderText(option.get('input_prompt', ''))
                    input_field.setEnabled(False)
                    
                    # 체크박스 상태에 따라 입력 필드 활성화/비활성화
                    def update_input_state(state, input_widget):
                        input_widget.setEnabled(state == Qt.CheckState.Checked)
                        if not state == Qt.CheckState.Checked:
                            input_widget.clear()
                    
                    checkbox.stateChanged.connect(lambda state, field=input_field: update_input_state(state, field))
                    
                    option_layout.addWidget(checkbox)
                    option_layout.addWidget(input_field)
                    option_widget.setLayout(option_layout)
                    options_layout.addWidget(option_widget)
                else:
                    checkbox = QCheckBox(option['label'])
                    checkbox.setProperty('value', option['value'])
                    options_layout.addWidget(checkbox)
            
            options_widget.show()
        else:
            dialog = item.treeWidget().window()
            options_widget = dialog.findChild(QWidget, "options_widget")
            if options_widget:
                options_widget.hide()
    except Exception as e:
        logger.error(f"옵션 UI 표시 중 오류 발생: {e}")
        QMessageBox.critical(None, "오류", f"옵션 UI 표시 중 오류가 발생했습니다: {e}")

def add_to_favorites(command_name, favorites, settings, update_ui_callback):
    """즐겨찾기에 명령어를 추가합니다."""
    try:
        if command_name not in favorites:
            favorites.append(command_name)
            settings.setValue('ssh_favorites', favorites)
            settings.sync()
            if update_ui_callback:
                update_ui_callback()
    except Exception as e:
        logger.error(f"즐겨찾기 추가 중 오류 발생: {e}")
        QMessageBox.critical(None, "오류", f"즐겨찾기 추가 중 오류가 발생했습니다: {e}")

def remove_from_favorites(command_name, favorites, settings, update_ui_callback):
    """즐겨찾기에서 명령어를 제거합니다."""
    try:
        if command_name in favorites:
            favorites.remove(command_name)
            settings.setValue('ssh_favorites', favorites)
            settings.sync()
            if update_ui_callback:
                update_ui_callback()
    except Exception as e:
        logger.error(f"즐겨찾기 제거 중 오류 발생: {e}")
        QMessageBox.critical(None, "오류", f"즐겨찾기 제거 중 오류가 발생했습니다: {e}")

def handle_favorite_click(dialog, command_map, command_name):
    """즐겨찾기 항목 클릭 처리"""
    try:
        # command_map에서 해당 명령어 찾기
        for category in command_map.values():
            for cmd_name, cmd_info in category.items():
                if cmd_name == command_name:
                    # 트리에서 해당 항목 선택
                    root = dialog.findChild(QTreeWidget)
                    if root:
                        # 모든 최상위 항목을 순회
                        for i in range(root.topLevelItemCount()):
                            category_item = root.topLevelItem(i)
                            # 카테고리의 모든 자식 항목을 순회
                            for j in range(category_item.childCount()):
                                command_item = category_item.child(j)
                                if command_item.text(0) == command_name:
                                    root.setCurrentItem(command_item)
                                    # 옵션이 있는 경우 옵션 UI 표시
                                    if cmd_info.get('has_options'):
                                        show_options(command_item)
                                    return
    except Exception as e:
        logger.error(f"즐겨찾기 클릭 처리 중 오류 발생: {e}")
        QMessageBox.critical(dialog, "오류", f"즐겨찾기 처리 중 오류가 발생했습니다: {e}")

def show_ssh_command_dialog(parent):
    dialog = QDialog(parent)
    dialog.setWindowTitle("SSH 명령어 선택")
    dialog.resize(400, 300)  # 대화상자 크기 조정
    
    # 레이아웃 설정
    layout = QVBoxLayout()
    
    # 명령어와 설명을 매핑하는 딕셔너리
    command_map = {
        "호스트네임 관리": {
            "호스트네임 조회": {
                "command": "racadm get system.serveros.hostname",
                "needs_input": False
            },
            "호스트네임 변경": {
                "command": "racadm set system.serveros.hostname '{hostname}'",
                "needs_input": True
            }
        },
        "시스템 로그 관리": {
            "SEL 로그 초기화": {
                "command": "racadm clrsel",
                "needs_input": False,
                "needs_confirm": True,
                "confirm_message": "SEL 로그를 초기화하시겠습니까?",
                "post_action": "refresh_sel"
            },
            "TSR 로그 수집": {  # TSR 로그 수집 명령어 추가
                "command": "racadm techsupreport collect -t Sysinfo,TTYLog",
                "needs_input": False,
                "needs_confirm": False,
                "is_tsr": True  # TSR 로그 수집임을 표시
            }
        },
        "카페24": {
            "카페24 관리": {
                "command": "",  # 실제 명령어는 선택된 옵션에 따라 동적으로 생성
                "needs_input": False,
                "has_options": True,
                "options": [
                    {"label": "패스워드 정책 확인", "value": "check_policy"},
                    {"label": "패스워드 기본값으로 변경 (default = calvin)", "value": "option1"},
                    {"label": "패스워드 직접 입력하여 변경", "value": "option1_custom", "needs_input": True, "input_prompt": "새로운 패스워드를 입력하세요"},
                    {"label": "전체 설정 조회 (논리 프로세서/BIOS/프로파일)", "value": "check_all"},
                    {"label": "전체 설정 변경 (Disabled/BIOS/Performance)", "value": "set_all"},
                    {"label": "논리 프로세서 설정 조회", "value": "option2"},
                    {"label": "논리 프로세서 설정 Disabled로 변경", "value": "option2_set"},
                    {"label": "BIOS 부트 모드 조회", "value": "option3"},
                    {"label": "BIOS 모드로 변경", "value": "option3_set"},
                    {"label": "프로파일 설정 조회", "value": "option4"},
                    {"label": "프로파일 설정 Performance로 변경", "value": "option4_set"},
                    {"label": "BIOS 설정 적용을 위한 시스템 재시작", "value": "option5"}
                ]
            }
        }
    }
    
    # 설정 로드
    settings = QSettings('Dell', 'iDRAC Monitor')
    favorites = settings.value('ssh_favorites', [], type=list)
    
    # 즐겨찾기 섹션과 그룹박스를 클래스 변수로 저장
    dialog.favorites = favorites
    dialog.favorites_group = None
    
    def update_favorites_ui():
        # 기존 즐겨찾기 그룹박스가 있다면 제거
        if dialog.favorites_group:
            layout.removeWidget(dialog.favorites_group)
            dialog.favorites_group.deleteLater()
            dialog.favorites_group = None
        
        # 즐겨찾기가 있는 경우에만 그룹박스 생성
        if dialog.favorites:
            dialog.favorites_group = QGroupBox("즐겨찾기")
            favorites_layout = QVBoxLayout()
            for fav in dialog.favorites:
                btn = QPushButton(fav)
                btn.clicked.connect(lambda checked, cmd=fav: handle_favorite_click(dialog, command_map, cmd))
                favorites_layout.addWidget(btn)
            dialog.favorites_group.setLayout(favorites_layout)
            layout.insertWidget(0, dialog.favorites_group)  # 항상 최상단에 추가
    
    # 초기 즐겨찾기 UI 생성
    update_favorites_ui()
    
    # 트리 위젯 생성
    tree = QTreeWidget()
    tree.setHeaderLabels(["명령어"])
    tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
    
    def handle_context_menu(pos):
        item = tree.itemAt(pos)
        if not item:
            return
            
        command_info = item.data(0, Qt.ItemDataRole.UserRole)
        if not command_info:  # 카테고리가 선택된 경우
            return
            
        menu = QMenu()
        command_name = item.text(0)
        
        if command_name in dialog.favorites:
            action = menu.addAction("즐겨찾기 제거")
            action.triggered.connect(lambda: remove_from_favorites(command_name, dialog.favorites, settings, update_favorites_ui))
        else:
            action = menu.addAction("즐겨찾기 추가")
            action.triggered.connect(lambda: add_to_favorites(command_name, dialog.favorites, settings, update_favorites_ui))
        
        menu.exec(tree.viewport().mapToGlobal(pos))
    
    tree.customContextMenuRequested.connect(handle_context_menu)
    
    # 더블클릭 이벤트 추가
    tree.itemDoubleClicked.connect(lambda: dialog.accept())
    
    # 트리 아이템 추가
    for category, commands in command_map.items():
        category_item = QTreeWidgetItem([category])
        for cmd_name, cmd_info in commands.items():
            command_item = QTreeWidgetItem([cmd_name])
            command_item.setData(0, Qt.ItemDataRole.UserRole, cmd_info)
            category_item.addChild(command_item)
        tree.addTopLevelItem(category_item)
    
    tree.expandAll()
    layout.addWidget(tree)
    
    # 옵션 체크박스를 저장할 위젯
    options_widget = QWidget()
    options_widget.setObjectName("options_widget")  # 위젯 이름 설정
    options_layout = QVBoxLayout()
    options_widget.setLayout(options_layout)
    options_widget.hide()
    layout.addWidget(options_widget)
        
    # 아이템 선택 변경 이벤트 연결
    tree.itemClicked.connect(show_options)
    tree.itemDoubleClicked.connect(show_options)
    
    # 확인/취소 버튼
    button_box = QDialogButtonBox(
        QDialogButtonBox.StandardButton.Ok | 
        QDialogButtonBox.StandardButton.Cancel
    )
    button_box.accepted.connect(dialog.accept)
    button_box.rejected.connect(dialog.reject)
    layout.addWidget(button_box)
    
    dialog.setLayout(layout)
    
    # 대화상자 실행 및 결과 반환
    result = dialog.exec()
    
    # Cancel 버튼을 눌렀으면 종료
    if result == QDialog.DialogCode.Rejected:
        return False, None
    
    # OK 버튼을 눌렀으면
    selected_items = tree.selectedItems()
    if not selected_items:
        return True, None
        
    selected_item = selected_items[0]
    command_info = selected_item.data(0, Qt.ItemDataRole.UserRole)
    
    if not command_info:  # 카테고리가 선택된 경우
        return True, None
        
    # 카페24 명령어 처리
    if command_info.get('has_options'):
        selected_options = []
        custom_inputs = {}
        
        for i in range(options_layout.count()):
            widget = options_layout.itemAt(i).widget()
            if isinstance(widget, QWidget) and widget.layout():
                # 입력 필드가 있는 옵션의 경우
                h_layout = widget.layout()
                checkbox = None
                input_field = None
                
                # 수평 레이아웃에서 체크박스와 입력 필드 찾기
                for j in range(h_layout.count()):
                    item = h_layout.itemAt(j).widget()
                    if isinstance(item, QCheckBox):
                        checkbox = item
                    elif isinstance(item, QLineEdit):
                        input_field = item
                
                if checkbox and checkbox.isChecked():
                    value = checkbox.property('value')
                    selected_options.append(value)
                    if input_field and input_field.text():
                        custom_inputs[value] = input_field.text()
            
            elif isinstance(widget, QCheckBox):
                # 일반 체크박스의 경우
                if widget.isChecked():
                    selected_options.append(widget.property('value'))
        
        if selected_options:
            command_info = command_info.copy()
            command_info['command'] = cafe24_manager.execute_command(selected_options, custom_inputs)
        
    return True, command_info

def collect_tsr_log(parent, host, username, password=None):
    """TSR 로그를 Redfish API를 통해 수집하고 로컬로 다운로드합니다."""
    progress = QProgressDialog("TSR 로그 수집 중...", "취소", 0, 100, parent)
    progress.setWindowModality(Qt.WindowModality.WindowModal)
    progress.setAutoClose(True)
    progress.setAutoReset(True)
    progress.setMinimumDuration(0)
    progress.show()

    # 홈 디렉토리에 저장할 파일명 생성
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    tsr_filename = f"tsr_log_{host}_{timestamp}.zip"
    home_dir = str(Path.home())
    local_path = os.path.join(home_dir, "Downloads", tsr_filename)

    def update_progress():
        nonlocal progress_value
        if progress_value < 95:  # 95%까지만 자동으로 증가
            progress_value += 1
            progress.setValue(progress_value)

    try:
        progress_value = 0
        progress.setValue(progress_value)
        progress.setLabelText("TSR 로그 수집 중...")

        # Redfish API 엔드포인트
        base_url = f"https://{host}/redfish/v1"
        managers_url = f"{base_url}/Managers/iDRAC.Embedded.1"
        export_url = f"{managers_url}/Oem/Dell/DellLCService/Actions/DellLCService.ExportTechSupportReport"

        # 인증 및 헤더 설정
        auth = (username, password) if password else None
        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }

        # TSR 수집 요청
        data = {
            "ShareType": "Local",
            "DataSelectorArrayIn": ["SelLog", "TTYLog"],
            "FileName": tsr_filename
        }

        # SSL 검증 비활성화 (자체 서명 인증서 사용 시)
        import urllib3
        urllib3.disable_warnings()

        # 진행 상태 업데이트를 위한 타이머 설정
        timer = QTimer(parent)
        timer.timeout.connect(update_progress)
        timer.start(1000)  # 1초마다 업데이트

        # TSR 수집 요청 보내기
        response = requests.post(
            export_url,
            json=data,
            auth=auth,
            headers=headers,
            verify=False  # SSL 검증 비활성화
        )

        if response.status_code != 202:
            raise Exception(f"TSR 로그 수집 요청 실패: {response.text}")

        # 작업 상태 모니터링
        task_uri = response.headers.get('Location')
        if not task_uri:
            task_uri = response.json().get('@odata.id')

        if not task_uri:
            raise Exception("작업 상태를 모니터링할 수 없습니다.")

        # 작업 완료 대기
        while True:
            task_response = requests.get(
                f"https://{host}{task_uri}",
                auth=auth,
                headers=headers,
                verify=False
            )
            
            task_data = task_response.json()
            if task_data.get('TaskState') == 'Completed':
                break
            elif task_data.get('TaskState') in ['Failed', 'Exception', 'Killed']:
                raise Exception(f"TSR 로그 수집 실패: {task_data.get('Messages', [{}])[0].get('Message')}")
            
            time.sleep(2)

        progress.setValue(100)
        QMessageBox.information(parent, "완료", f"TSR 로그가 성공적으로 수집되었습니다.\n저장 위치: {local_path}")

    except Exception as e:
        logger.error(f"TSR 로그 수집 중 오류 발생: {str(e)}")
        QMessageBox.critical(parent, "오류", f"TSR 로그 수집 중 오류가 발생했습니다: {str(e)}")
    finally:
        timer.stop() if 'timer' in locals() else None
        progress.close()

def open_ssh_connection(parent):
    try:
        main_window = parent.window()
        if not hasattr(main_window, 'server_section'):
            error_dialog = ErrorDialog(
                "서버 연결 오류",
                "서버가 연결되어 있지 않습니다.",
                "서버를 먼저 연결한 후 다시 시도해주세요.",
                parent
            )
            error_dialog.exec()
            return
            
        server_info = main_window.server_section.current_server_info
        if not server_info:
            error_dialog = ErrorDialog(
                "서버 연결 오류",
                "서버 정보를 찾을 수 없습니다.",
                "서버를 선택한 후 다시 시도해주세요.",
                parent
            )
            error_dialog.exec()
            return
        
        # SSH 명령어 선택 대화상자 표시
        proceed, command_info = show_ssh_command_dialog(parent)
        
        # Cancel 버튼을 눌렀으면 종료
        if not proceed or not command_info:
            return
            
        from utils.ssh_utils import open_ssh_terminal
        
        # SSH 연결 시도
        ssh_params = {
            "host": server_info['IP'],
            "username": 'root',
            "key_path": '~/.ssh/id_rsa',
            "password": server_info.get('PASSWORD')
        }
        
        # TSR 로그 수집인 경우
        if command_info.get('is_tsr'):
            collect_tsr_log(parent, ssh_params['host'], ssh_params['username'], ssh_params['password'])
        else:
            # 일반 SSH 명령어 실행
            if command_info.get('command') is not None:
                ssh_params["command"] = command_info['command']
            
            # SSH 명령어 실행
            open_ssh_terminal(**ssh_params)
        
            # SEL 로그 초기화 후 갱신
            if command_info.get('command') == "racadm clrsel":
                QTimer.singleShot(2000, lambda: refresh_sel_after_clear(main_window))
            
    except Exception as e:
        logger.error(f"SSH 연결 실패: {str(e)}")
        error_dialog = ErrorDialog(
            "SSH 연결 오류",
            "SSH 연결 중 오류가 발생했습니다.",
            str(e),
            parent
        )
        error_dialog.exec()

def show_all_status(parent):
    """CPU(GPU포함) 정보 / MEMORY 정보 / STORAGE 정보 / NIC 정보 / PSU 정보 / iDRAC MAC 정보를 통합하여 테이블 형식으로 반환"""
    logger.debug("시스템 상태 정보 조회 시도")
    
    main_window = parent.window()
    if not hasattr(main_window, 'server_section'):
        logger.warning("서버 섹션을 찾을 수 없음")
        error_dialog = ErrorDialog(
            "서버 연결 오류",
            "서버가 연결되어 있지 않습니다.",
            "서버를 먼저 연결한 후 다시 시도해주세요.",
            parent
        )
        error_dialog.exec()
        return
    
    server_info = main_window.server_section.current_server_info
    if not server_info:
        logger.warning("서버 정보가 없음")
        error_dialog = ErrorDialog(
            "서버 연결 오류",
            "서버 정보를 찾을 수 없습니다.",
            "서버를 선택한 후 다시 시도해주세요.",
            parent
        )
        error_dialog.exec()
        return
    
    # 진행률 다이얼로그 생성
    progress_dialog = QProgressDialog("시스템 상태 정보 로딩 중...", None, 0, 100, parent)
    progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
    progress_dialog.setWindowTitle("데이터 로드")
    progress_dialog.setAutoClose(True)
    progress_dialog.setMinimumDuration(0)
    
    status_dialog = QDialog(parent)
    status_dialog.setWindowTitle("System Information")
    status_dialog.resize(800, 600)

    layout = QVBoxLayout()
    status_dialog.setLayout(layout)

    button_layout = QHBoxLayout()
    expand_collapse_button = QPushButton("전체 펼치기")
    rebuild_monitor_toggle = QPushButton("리빌딩 모니터링")
    rebuild_monitor_toggle.setCheckable(True)
    rebuild_monitor_toggle.setStyleSheet("""
        QPushButton:checked {
            background-color: #FFA500;
            color: white;
        }
    """)
    
    button_layout.addWidget(expand_collapse_button)
    button_layout.addWidget(rebuild_monitor_toggle)
    layout.addLayout(button_layout)

    tree_widget = QTreeWidget()
    tree_widget.setHeaderLabels(["구성 요소", "Dell Attribute name", "value"])
    tree_widget.setColumnWidth(0, 250)
    tree_widget.setColumnWidth(1, 250)
    layout.addWidget(tree_widget)

    # 리빌딩 상태를 표시할 영역
    rebuild_status_layout = QVBoxLayout()
    rebuild_status_group = QGroupBox("리빌딩 상태")
    rebuild_status_group.setVisible(False)
    rebuild_status_tree = QTreeWidget()
    rebuild_status_tree.setHeaderLabels(["디스크 위치", "진행률", "예상 시간"])
    rebuild_status_layout.addWidget(rebuild_status_tree)
    rebuild_status_group.setLayout(rebuild_status_layout)
    layout.addWidget(rebuild_status_group)

    def check_disk_status(drive):
        status = drive.get('Oem', {}).get('Dell', {}).get('DellPhysicalDisk', {}).get('RaidStatus', 'N/A')
        
        status_colors = {
            'Online': QColor('green'),
            'Rebuilding': QColor('orange'),
            'Failed': QColor('red'),
            'Degraded': QColor('yellow'),
            'Offline': QColor('gray')
        }
        
        return (status, status_colors.get(status))

    def update_rebuild_status():
        rebuild_status_tree.clear()
        storage_info = server_manager.fetch_storage_info()
        rebuilding_exists = False
        
        for controller in storage_info.get('Controllers', []):
            volumes = controller.get('Volumes', [])
            drives = controller.get('Drives', [])
            
            for volume in volumes:
                volume_name = volume.get('Name', '')
                raid_type = volume.get('RAIDType', '')
                capacity = volume.get('CapacityBytes', 0) / (1024**4)
                
                volume_drive_ids = [link.get('@odata.id', '').split('/')[-1] 
                                for link in volume.get('Links', {}).get('Drives', [])]
                
                volume_item = QTreeWidgetItem(rebuild_status_tree)
                volume_item.setText(0, f"볼륨: {volume_name}")
                volume_item.setText(1, f"RAID {raid_type}")
                volume_item.setText(2, f"{capacity:.1f} TiB")
                volume_item.setBackground(0, QColor('#E6E6FA'))  # 연한 녹색 배경
                
                # 리빌딩 디스크 존재 여부 확인을 위한 플래그
                has_rebuilding_disk = False
                
                for drive in drives:
                    if drive.get('Id', '') in volume_drive_ids:
                        disk_id = drive.get('Id', '').split(':')[0]
                        
                        drive_item = QTreeWidgetItem(volume_item)
                        drive_item.setText(0, disk_id)
                        
                        status, color = check_disk_status(drive)
                        
                        if drive.get('Operations'):
                            for operation in drive.get('Operations', []):
                                if operation.get('OperationName') == "Rebuilding":
                                    rebuilding_exists = True
                                    has_rebuilding_disk = True  # 리빌딩 중인 디스크 발견
                                    progress = operation.get('PercentageComplete', 0)
                                    
                                    progress_bar = QProgressBar()
                                    progress_bar.setValue(progress)
                                    progress_bar.setStyleSheet("""
                                        QProgressBar {
                                            border: 2px solid grey;
                                            text-align: center;
                                        }
                                        QProgressBar::chunk {
                                            background-color: #FFA500;
                                        }
                                    """)
                                    rebuild_status_tree.setItemWidget(drive_item, 1, progress_bar)
                                    
                                    if progress > 0:
                                        remaining_minutes = int((100 - progress) * 2)
                                        drive_item.setText(2, f"{remaining_minutes}분")
                        else:
                            status_label = QLabel(status)
                            if color:
                                status_label.setStyleSheet(f"color: {color.name()}")
                            rebuild_status_tree.setItemWidget(drive_item, 1, status_label)
                            drive_item.setText(2, "-")
                
                # 볼륨의 펼침 상태 설정
                if has_rebuilding_disk:
                    rebuild_status_tree.expandItem(volume_item)  # 리빌딩 중인 디스크가 있는 볼륨은 펼치기
                else:
                    rebuild_status_tree.collapseItem(volume_item)  # 리빌딩 중인 디스크가 없는 볼륨은 접기
        
        if not rebuilding_exists:
            rebuild_monitor_toggle.setChecked(False)
            rebuild_status_group.setVisible(False)
    
    def toggle_rebuild_monitor(checked):
        rebuild_status_group.setVisible(checked)
        if checked:
            update_rebuild_status()
            timer.start(10000)  # 10초마다 갱신
        else:
            timer.stop()
    
    timer = QTimer()
    timer.timeout.connect(update_rebuild_status)
    rebuild_monitor_toggle.toggled.connect(toggle_rebuild_monitor)
    
    try:
        def toggle_all_sections():
            if expand_collapse_button.text() == "전체 펼치기":
                tree_widget.expandAll()
                expand_collapse_button.setText("전체 접기")
            else:
                tree_widget.collapseAll()
                expand_collapse_button.setText("전체 펼치기")
        
        expand_collapse_button.clicked.connect(toggle_all_sections)

        main_window = parent.window()
        if hasattr(main_window, 'server_section'):
            server_info = main_window.server_section.current_server_info
            if server_info:
                server_manager = DellServerManager(
                    ip=server_info['IP'],
                    port=server_info['PORT'],
                    auth=(server_info['USERNAME'], server_info['PASSWORD'])
                )
                
                progress_dialog.show()

                # 데이터 로드
                data = {
                    'processors': server_manager.fetch_processors_info(),
                    'memory': server_manager.fetch_memory_info(),
                    'storage': server_manager.fetch_storage_info(),
                    'nic': server_manager.fetch_network_adapters_info(),
                    'psu': server_manager.fetch_psu_info(),
                    'idrac': server_manager.fetch_detailed_info(server_manager.endpoints.idrac_mac_address),
                    'license': server_manager.check_idrac_license()
                }

                # 섹션별 설정 딕셔너리 정의
                processor_settings = {
                    "모델": "Model",
                    "제조사": "Manufacturer",
                    "코어 수": "TotalCores",
                    "스레드 수": "TotalThreads",
                    "최대 속도": "MaxSpeedMHz",
                    "현재 속도": "OperatingSpeedMHz",
                    "상태": "Status.Health"
                }
                dell_processor_settings = {
                    "하이퍼스레딩": "HyperThreadingEnabled",
                    "가상화 기술": "VirtualizationTechnologyEnabled",
                    "터보 모드": "TurboModeEnabled"
                }
                memory_settings = {
                    "제조사": "Manufacturer",
                    "타입": "MemoryDeviceType",
                    "동작 속도": "OperatingSpeedMhz",
                    "용량": "CapacityMiB",
                    "상태": "Status.Health"
                }
                storage_settings = {
                    "모델": "Model",
                    "펌웨어 버전": "FirmwareVersion",
                    "캐시": "TotalCacheSizeMiB",
                    "상태": "Status.Health"
                }
                volume_settings = {
                    "레이드": "RAIDType",
                    "미디어 타입": "MediaType",
                    "용량": "CapacityBytes",
                    "상태": "RaidStatus"
                }
                drive_settings = {
                    "제조사": "Manufacturer",
                    "파트 번호": "PartNumber",
                    "시리얼 번호": "SerialNumber",
                    "용량": "CapacityBytes",
                    "레이드 상태": "RaidStatus"
                }
                nic_settings = {
                    "모델": "Model",
                    "제조사": "Manufacturer",
                    "파트 번호": "PartNumber",
                    "시리얼 번호": "SerialNumber",
                    "상태": "Status.Health"
                }
                controller_settings = {
                    "펌웨어 버전": "FirmwarePackageVersion",
                    "가상화 지원": "VirtualizationOffload"
                }
                port_settings = {
                    "링크 상태": "LinkStatus",
                    "현재 속도": "CurrentLinkSpeedMbps",
                    "Flow Control 설정": "FlowControlConfiguration",
                    "Flow Control 상태": "FlowControlStatus",
                    "MAC 주소": "AssociatedNetworkAddresses"
                }
                transceiver_settings = {
                    "트랜시버 타입": "IdentifierType",
                    "인터페이스": "InterfaceType",
                    "트랜시버 제조사": "VendorName"
                }
                optical_settings = {
                    "온도": "Temperature",
                    "전압": "SupplyVoltage",
                    "TX 파워": "TxPower",
                    "RX 파워": "RxPower",
                    "레이저 바이어스 전류": "LaserBiasCurrent"
                }
                psu_settings = {
                    "모델": "Model",
                    "제조사": "Manufacturer",
                    "용량": "PowerCapacityWatts",
                    "상태": "Status.Health",
                    "펌웨어 버전": "FirmwareVersion",
                    "시리얼 번호": "SerialNumber",
                    "파트 번호": "PartNumber",
                }

                idrac_mac_settings = {
                    "MAC 주소": "CurrentNIC.1.MACAddress"
                }

                # 섹션 정의
                sections = [
                    ("프로세서 정보", data['processors'], processor_settings),
                    ("메모리 정보", data['memory'], memory_settings),
                    ("스토리지 정보", data['storage'], storage_settings),
                    ("NIC 정보", data['nic'], nic_settings),
                    ("PSU 정보", data['psu'], psu_settings),
                    ("iDRAC MAC 주소 정보", data['idrac'], idrac_mac_settings)
                ]

                # 섹션별 트리 아이템 생성
                for section_name, info_source, settings_dict in sections:
                    if info_source:
                        section_item = QTreeWidgetItem(tree_widget, [section_name])
                        
                        if section_name == "프로세서 정보":
                            if 'Members' in info_source:
                                for cpu in info_source['Members']:
                                    member_uri = cpu.get('@odata.id')
                                    if member_uri:
                                        cpu_response = requests.get(
                                            f"{server_manager.endpoints.base_url}{member_uri}",
                                            auth=server_manager.auth,
                                            verify=False
                                        )
                                        cpu_info = cpu_response.json()
                                        cpu_item = QTreeWidgetItem(section_item, [f"CPU {cpu_info.get('Id', 'N/A')}"])
                                        
                                        for key, value in settings_dict.items():
                                            item = QTreeWidgetItem(cpu_item)
                                            item.setText(0, key)
                                            item.setText(1, value)
                                            item.setText(2, str(cpu_info.get(value, 'N/A')))
                                            
                                            if key == "상태":
                                                status = cpu_info.get('Status', {})
                                                state = status.get('State')
                                                health = status.get('Health')
                                                
                                                if state == 'Enabled' and health == 'OK':
                                                    item.setText(2, 'OK')
                                                    item.setForeground(2, QColor('green'))
                                                elif state == 'Enabled' and health == 'Critical':
                                                    item.setText(2, 'Critical')
                                                    item.setForeground(2, QColor('red'))
                                                else:
                                                    item.setText(2, str(health))
                                        
                                        dell_info = cpu_info.get('Oem', {}).get('Dell', {}).get('DellProcessor', {})
                                        if dell_info:
                                            dell_section = QTreeWidgetItem(cpu_item, ["Dell 특정 정보"])
                                            for key, value in dell_processor_settings.items():
                                                item = QTreeWidgetItem(dell_section)
                                                item.setText(0, key)
                                                item.setText(1, value)
                                                enabled = "활성화" if dell_info.get(value) == "Yes" else "비활성화"
                                                item.setText(2, enabled)
                                                item.setForeground(2, QColor('green') if enabled == "활성화" else QColor('red'))
                        elif section_name == "메모리 정보":
                            if 'Members' in info_source:
                                sorted_members = sorted(info_source.get('Members', []), 
                                                    key=lambda x: x.get('@odata.id', ''))
                                for member in sorted_members:
                                    member_uri = member.get('@odata.id')
                                    if member_uri:
                                        memory_response = requests.get(
                                            f"{server_manager.endpoints.base_url}{member_uri}",
                                            auth=server_manager.auth,
                                            verify=False
                                        )
                                        memory_info = memory_response.json()
                                        
                                        memory_item = QTreeWidgetItem(section_item, 
                                                                    [f"메모리 {memory_info.get('Id', 'N/A')}"])
                                        
                                        for key, value in settings_dict.items():
                                            item = QTreeWidgetItem(memory_item)
                                            item.setText(0, key)
                                            item.setText(1, value)
                                            
                                            if key == "용량":
                                                value = convert_capacity(memory_info.get('CapacityMiB', 0), False)
                                            elif key == "동작 속도":
                                                value = f"{memory_info.get('OperatingSpeedMhz', 'N/A')} MHz"
                                            else:
                                                value = memory_info.get(value, 'N/A')
                                            
                                            item.setText(2, str(value))
                                            
                                            if key == "상태":
                                                status = memory_info.get('Status', {})
                                                state = status.get('State')
                                                health = status.get('Health')
                                                
                                                if state == 'Enabled' and health == 'OK':
                                                    item.setText(2, 'OK')
                                                    item.setForeground(2, QColor('green'))
                                                elif state == 'Enabled' and health == 'Critical':
                                                    item.setText(2, 'Critical')
                                                    item.setForeground(2, QColor('red'))
                                                else:
                                                    item.setText(2, str(health))

                        elif section_name == "스토리지 정보":
                            if 'Controllers' in info_source:
                                for controller in info_source['Controllers']:
                                    storage_controllers = controller.get('StorageControllers', [])
                                    if storage_controllers:
                                        controller_info = storage_controllers[0]
                                        controller_item = QTreeWidgetItem(section_item,
                                            [f"컨트롤러 {controller.get('Id', 'N/A')}"])
                                        
                                        # 컨트롤러 정보 표시
                                        for key, value in storage_settings.items():
                                            item = QTreeWidgetItem(controller_item)
                                            item.setText(0, key)
                                            item.setText(1, value)
                                            
                                            if key == "캐시":
                                                cache_value = controller_info.get('CacheSummary', {}).get('TotalCacheSizeMiB', 0)
                                                item.setText(2, convert_capacity(cache_value, False))
                                            elif key == "상태":
                                                status = controller.get('Status', {}).get('Health', 'N/A')
                                                item.setText(2, str(status))
                                                if status == 'OK':
                                                    item.setForeground(2, QColor('green'))
                                            else:
                                                item.setText(2, str(controller_info.get(value, 'N/A')))

                                        # 볼륨 정보 표시
                                        volumes = controller.get('Volumes', [])
                                        for volume in volumes:
                                            dell_volume = volume.get('Oem', {}).get('Dell', {}).get('DellVolume', {})
                                            volume_item = QTreeWidgetItem(controller_item,
                                                [f"볼륨: {volume.get('Name', 'N/A')}"])
                                            
                                            for key, value in volume_settings.items():
                                                item = QTreeWidgetItem(volume_item)
                                                item.setText(0, key)
                                                item.setText(1, value)
                                                
                                                if key == "용량":
                                                    item.setText(2, convert_capacity(volume.get('CapacityBytes', 0), True))
                                                elif key == "미디어 타입":
                                                    item.setText(2, str(dell_volume.get('MediaType', 'N/A')))
                                                elif key == "상태":
                                                    status = dell_volume.get('RaidStatus', 'N/A')
                                                    item.setText(2, str(status))
                                                    if status == 'Online':
                                                        item.setForeground(2, QColor('green'))
                                                else:
                                                    item.setText(2, str(volume.get(value, 'N/A')))

                                            # 드라이브 정보 표시
                                            drives = controller.get('Drives', [])
                                            volume_drive_ids = [link.get('@odata.id', '').split('/')[-1]
                                                                for link in volume.get('Links', {}).get('Drives', [])]
                                            volume_drives = [d for d in drives if d.get('Id', '') in volume_drive_ids]
                                            sorted_drives = sort_drives(volume_drives)

                                            for drive in sorted_drives:
                                                simplified_id = drive.get('Id', 'N/A').split(':')[0]  # drive 변수가 정의된 후에 사용
                                                drive_item = QTreeWidgetItem(volume_item, [f"드라이브: {simplified_id}"])
                                                
                                                for key, value in drive_settings.items():
                                                    item = QTreeWidgetItem(drive_item)
                                                    item.setText(0, key)
                                                    item.setText(1, value)
                                                    
                                                    if key == "용량":
                                                        item.setText(2, convert_capacity(drive.get('CapacityBytes', 0), True))
                                                    elif key == "레이드 상태":
                                                        status = drive.get('Oem', {}).get('Dell', {}).get('DellPhysicalDisk', {}).get('RaidStatus', 'N/A')
                                                        item.setText(2, str(status))
                                                        if status == 'Online':
                                                            item.setForeground(2, QColor('green'))
                                                        elif status == 'Rebuilding':
                                                            item.setForeground(2, QColor('orange'))
                                                    else:
                                                        item.setText(2, str(drive.get(value, 'N/A')))
                                                
                                                # 리빌딩 감지 및 모니터링 버튼 추가
                                                if drive.get('Operations'):
                                                    for operation in drive.get('Operations', []):
                                                        if operation.get('OperationName') == "Rebuilding":
                                                            # 리빌딩 상태 표시
                                                            rebuild_status = QTreeWidgetItem(drive_item)
                                                            rebuild_status.setText(0, "리빌딩 상태")
                                                            rebuild_status.setText(1, "RebuildStatus")
                                                            rebuild_status.setText(2, "진행 중")
                                                            rebuild_status.setForeground(2, QColor('orange'))
                                                            
                                                            # 진행률 표시
                                                            progress = operation.get('PercentageComplete', 0)
                                                            progress_item = QTreeWidgetItem(drive_item)
                                                            progress_item.setText(0, "진행률")
                                                            progress_item.setText(1, "PercentageComplete")
                                                            progress_item.setText(2, f"{progress}%")
                                                            
                                                            # 모니터링 버튼 추가
                                                            monitor_button = QPushButton("리빌딩 모니터링")
                                                            monitor_button.setStyleSheet("background-color: #FFA500; color: white;")

                                                            # 현재 드라이버의 리빌딩 상태에 따른 모니터링
                                                            def create_monitor_handler(current_drive):
                                                                def show_rebuild_monitor():
                                                                    monitor_dialog = QDialog(parent)
                                                                    monitor_dialog.setWindowTitle(f"리빌딩 모니터링 - 드라이브 {current_drive.get('Id', 'N/A')}")
                                                                    monitor_dialog.resize(400, 150)
                                                                    
                                                                    # 메인 레이아웃
                                                                    main_layout = QVBoxLayout()

                                                                    # 갱신 주기 설정
                                                                    refresh_layout = QHBoxLayout()
                                                                    refresh_label = QLabel("갱신 주기(초):")
                                                                    refresh_spin = QSpinBox()
                                                                    refresh_spin.setRange(5, 60)
                                                                    refresh_spin.setValue(10)
                                                                    refresh_spin.setToolTip("5초에서 60초 사이로 설정 가능합니다")
                                                                    refresh_layout.addWidget(refresh_label)
                                                                    refresh_layout.addWidget(refresh_spin)

                                                                    # 상태 표시 레이블 (진행률과 예상 시간)
                                                                    status_label = QLabel()
                                                                    status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                                                                    
                                                                    # 프로그레스바와 취소 버튼이 있는 영역
                                                                    progress_layout = QVBoxLayout()
                                                                    progress_bar = QProgressBar()
                                                                    progress_bar.setValue(progress)
                                                                    progress_bar.setStyleSheet("""
                                                                        QProgressBar {
                                                                            border: 2px solid grey;
                                                                            text-align: center;
                                                                        }
                                                                        QProgressBar::chunk {
                                                                            background-color: #FFA500;
                                                                        }
                                                                    """)
                                                                    cancel_button = QPushButton("모니터링 중지")
                                                                    progress_layout.addWidget(progress_bar)
                                                                    progress_layout.addWidget(cancel_button)

                                                                    # 레이아웃 구성
                                                                    main_layout.addLayout(refresh_layout)
                                                                    main_layout.addWidget(status_label)
                                                                    main_layout.addLayout(progress_layout)
                                                                    monitor_dialog.setLayout(main_layout)
                                                                    
                                                                    # 타이머 설정
                                                                    start_time = time.time()
                                                                    last_progress = progress
                                                                    timer = QTimer(parent)
                                                                    
                                                                    def update_progress():
                                                                        try:
                                                                            drive_info = server_manager.fetch_drive_rebuild_status(current_drive)
                                                                            if drive_info and 'Operations' in drive_info:
                                                                                for op in drive_info['Operations']:
                                                                                    if op.get('OperationName') == "Rebuilding":
                                                                                        current_progress = op.get('PercentageComplete', 0)
                                                                                        
                                                                                        # 예상 시간 계산
                                                                                        elapsed_time = time.time() - start_time
                                                                                        if current_progress > 0:
                                                                                            total_time = (elapsed_time * 100) / current_progress
                                                                                            remaining_time = total_time - elapsed_time
                                                                                            remaining_minutes = int(remaining_time // 60)
                                                                                            remaining_seconds = int(remaining_time % 60)
                                                                                            
                                                                                            status_label.setText(
                                                                                                f"리빌딩 진행률: {current_progress}%\n"
                                                                                                f"예상 남은 시간: {remaining_minutes}분 {remaining_seconds}초"
                                                                                            )
                                                                                        
                                                                                        progress_bar.setValue(current_progress)
                                                                                        last_progress = current_progress
                                                                                        
                                                                        except Exception as e:
                                                                            logger.error(f"리빌딩 상태 업데이트 실패: {str(e)}")
                                                                        
                                                                        if last_progress == 100:
                                                                            timer.stop()
                                                                            monitor_dialog.close()
                                                                    
                                                                    def on_canceled():
                                                                        timer.stop()
                                                                        monitor_dialog.close()
                                                                    
                                                                    timer.timeout.connect(update_progress)
                                                                    cancel_button.clicked.connect(on_canceled)
                                                                    refresh_spin.valueChanged.connect(lambda: timer.setInterval(refresh_spin.value() * 1000))
                                                                    
                                                                    timer.start(refresh_spin.value() * 1000)
                                                                    monitor_dialog.exec()
                                                                
                                                                return show_rebuild_monitor

                                                            monitor_button.clicked.connect(create_monitor_handler(drive))
                                                            button_widget = QTreeWidgetItem(drive_item)
                                                            tree_widget.setItemWidget(button_widget, 2, monitor_button)

                        elif section_name == "NIC 정보":
                            if 'NetworkAdapters' in info_source:
                                sorted_adapters = sorted(info_source['NetworkAdapters'], key=lambda x: get_nic_order(x.get('Id', '')))
                                for adapter in sorted_adapters:
                                    adapter_item = QTreeWidgetItem(section_item, [f"NIC 어댑터: {adapter.get('Id', 'N/A')}"])
                                    
                                    # NIC 기본 정보
                                    for key, value in nic_settings.items():
                                        item = QTreeWidgetItem(adapter_item)
                                        item.setText(0, key)
                                        item.setText(1, value)
                                        if key == "상태":
                                            status = adapter.get('Status', {})
                                            health = status.get('Health', 'N/A')
                                            item.setText(2, health)
                                            if health == 'OK':
                                                item.setForeground(2, QColor('green'))
                                        else:
                                            item.setText(2, str(adapter.get(value, 'N/A')))

                                    # 컨트롤러 정보
                                    for controller in adapter.get('Controllers', []):
                                        controller_item = QTreeWidgetItem(adapter_item, ["컨트롤러 정보"])
                                        is_virtualization_supported = controller.get('ControllerCapabilities', {}).get(
                                            'VirtualizationOffload', {}).get('SRIOV', {}).get('SRIOVVEPACapable', False)

                                        for key, value in controller_settings.items():
                                            item = QTreeWidgetItem(controller_item)
                                            item.setText(0, key)
                                            item.setText(1, value)
                                            if key == "가상화 지원":
                                                item.setText(2, "가상화 지원 카드" if is_virtualization_supported else "가상화 미지원 카드")
                                            else:
                                                item.setText(2, str(controller.get(value, 'N/A')))

                                    # 포트 정보
                                    for port in adapter.get('NetworkPorts', []):
                                        port_id = port.get('Id', 'N/A')
                                        device_function_id = f"{port_id}-1"
                                        port_item = QTreeWidgetItem(adapter_item, [f"포트: {port_id}"])

                                        # 가상화 모드 정보
                                        virtualization_mode = 'N/A'
                                        if is_virtualization_supported:
                                            try:
                                                device_function_info = requests.get(
                                                    f"{server_manager.endpoints.base_url}/redfish/v1/Chassis/System.Embedded.1/NetworkAdapters/{adapter.get('Id')}/NetworkDeviceFunctions/{device_function_id}/Oem/Dell/DellNetworkAttributes/{device_function_id}",
                                                    auth=server_manager.auth,
                                                    verify=False
                                                ).json()
                                                virtualization_mode = device_function_info.get('Attributes', {}).get('VirtualizationMode', 'N/A')
                                            except:
                                                pass

                                        # 포트 설정 표시
                                        for key, value in port_settings.items():
                                            item = QTreeWidgetItem(port_item)
                                            item.setText(0, key)
                                            item.setText(1, value)
                                            if key == "현재 속도":
                                                item.setText(2, f"{port.get(value, 'N/A')} Mbps")
                                            elif key == "MAC 주소":
                                                addresses = port.get(value, ['N/A'])
                                                item.setText(2, addresses[0] if addresses else 'N/A')
                                            elif key == "링크 상태":
                                                status = port.get(value, 'N/A')
                                                item.setText(2, status)
                                                if status == 'Up':
                                                    item.setForeground(2, QColor('green'))
                                                elif status == 'Down':
                                                    item.setForeground(2, QColor('red'))
                                            else:
                                                item.setText(2, str(port.get(value, 'N/A')))

                                        if is_virtualization_supported:
                                            virt_item = QTreeWidgetItem(port_item)
                                            virt_item.setText(0, "가상화 모드")
                                            virt_item.setText(1, "VirtualizationMode")
                                            virt_item.setText(2, virtualization_mode)

                                        # 트랜시버 정보
                                        transceiver = port.get('Oem', {}).get('Dell', {}).get('DellNetworkTransceiver', {})
                                        if transceiver and data.get('license') and 'enterprise' in data['license']['type'].lower():
                                            transceiver_item = QTreeWidgetItem(port_item, ["트랜시버 정보"])
                                            
                                            for key, value in transceiver_settings.items():
                                                item = QTreeWidgetItem(transceiver_item)
                                                item.setText(0, key)
                                                item.setText(1, value)
                                                item.setText(2, str(transceiver.get(value, 'N/A')))

                                            # 광 레벨 정보
                                            if 'datacenter' in data['license']['type'].lower():
                                                optical_data = transceiver.get('OpticalData', {})
                                                if optical_data:
                                                    optical_item = QTreeWidgetItem(transceiver_item, ["광 레벨 정보"])
                                                    
                                                    for key, value in optical_settings.items():
                                                        item = QTreeWidgetItem(optical_item)
                                                        item.setText(0, key)
                                                        item.setText(1, value)
                                                        if value in optical_data:
                                                            if value == "Temperature":
                                                                display_value = f"{optical_data[value]} °C"
                                                            elif value == "SupplyVoltage":
                                                                display_value = f"{optical_data[value]} V"
                                                            elif value in ["TxPower", "RxPower"]:
                                                                display_value = f"{optical_data[value]} dBm"
                                                            elif value == "LaserBiasCurrent":
                                                                display_value = f"{optical_data[value]} mA"
                                                            else:
                                                                display_value = str(optical_data[value])
                                                        else:
                                                            display_value = 'N/A'
                                                        item.setText(2, display_value)
                        
                        elif section_name == "PSU 정보":
                            # PSU 정보 추가
                            if info_source and 'PowerSupplies' in info_source:
                                for psu in info_source['PowerSupplies']:
                                    psu_id = f"PSU {psu.get('MemberId', 'N/A')}"
                                    psu_item = QTreeWidgetItem(section_item, [psu_id])
                                    
                                    for key, value in settings_dict.items():
                                        item = QTreeWidgetItem(psu_item)
                                        item.setText(0, key)
                                        item.setText(1, value)
                                        
                                        # 특별한 형식이 필요한 필드들 처리
                                        if key == "용량":
                                            item.setText(2, f"{psu.get(value, 'N/A')}W")
                                        elif key == "상태":
                                            status = psu.get('Status', {})
                                            health = status.get('Health', 'N/A')
                                            item.setText(2, str(health))
                                            if health == "OK":
                                                item.setForeground(2, QColor('green'))
                                            elif health == "Critical":
                                                item.setForeground(2, QColor('red'))
                                        else:
                                            item.setText(2, str(psu.get(value, 'N/A')))

                        elif section_name == "iDRAC MAC 주소 정보":
                            # iDRAC MAC 주소 정보 추가
                            if 'Attributes' in info_source:
                                mac_address = info_source.get('Attributes', {}).get('CurrentNIC.1.MACAddress', 'N/A')
                                item = QTreeWidgetItem(section_item)
                                item.setText(0, "MAC 주소")
                                item.setText(1, "CurrentNIC.1.MACAddress")
                                item.setText(2, str(mac_address))

                tree_widget.collapseAll()
                progress_dialog.setValue(100)

                def show_status_dialog():
                    progress_dialog.close()
                    status_dialog.exec()

                QTimer.singleShot(500, show_status_dialog)

    except Exception as e:
        progress_dialog.close()
        logger.error(f"시스템 상태 정보 조회/표시 실패: {str(e)}")
        logger.exception(e)
        error_dialog = ErrorDialog(
            "시스템 상태 조회 오류",
            "시스템 상태 정보를 조회하는 중 오류가 발생했습니다.",
            str(e),
            parent
        )
        error_dialog.exec()
                
def show_system_info(parent):
    """BIOS 정보 / iDRAC 정보 / 네트워크 설정 정보를 통합하여 테이블 형식으로 반환"""
    logger.debug("시스템 정보 조회 시도")
    
    main_window = parent.window()
    if not hasattr(main_window, 'server_section'):
        logger.warning("서버 섹션을 찾을 수 없음")
        error_dialog = ErrorDialog(
            "서버 연결 오류",
            "서버가 연결되어 있지 않습니다.",
            "서버를 먼저 연결한 후 다시 시도해주세요.",
            parent
        )
        error_dialog.exec()
        return
    
    server_info = main_window.server_section.current_server_info
    if not server_info:
        logger.warning("서버 정보가 없음")
        error_dialog = ErrorDialog(
            "서버 연결 오류",
            "서버 정보를 찾을 수 없습니다.",
            "서버를 선택한 후 다시 시도해주세요.",
            parent
        )
        error_dialog.exec()
        return
    
    # 진행률 다이얼로그 생성
    progress_dialog = QProgressDialog("시스템 정보 로딩 중...", None, 0, 100, parent)
    progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
    progress_dialog.setWindowTitle("데이터 로드")
    progress_dialog.setAutoClose(True)
    progress_dialog.setMinimumDuration(0)
    
    status_dialog = QDialog(parent)
    status_dialog.setWindowTitle("System Information")
    status_dialog.resize(800, 600)
    
    layout = QVBoxLayout()
    expand_collapse_button = QPushButton("전체 펼치기")
    button_layout = QHBoxLayout()
    button_layout.addWidget(expand_collapse_button)
    layout.addLayout(button_layout)
    tree_widget = QTreeWidget()
    tree_widget.setHeaderLabels(["Settings", "Dell Attribute name", "value"])
    tree_widget.setColumnWidth(0, 250)
    tree_widget.setColumnWidth(1, 250)
    layout.addWidget(tree_widget)
    status_dialog.setLayout(layout)
    
    try:
        def toggle_all_sections():
            if expand_collapse_button.text() == "전체 펼치기":
                tree_widget.expandAll()
                expand_collapse_button.setText("전체 접기")
            else:
                tree_widget.collapseAll()
                expand_collapse_button.setText("전체 펼치기")
        
        expand_collapse_button.clicked.connect(toggle_all_sections)
        
        # 진행률 다이얼로그 생성
        main_window = parent.window()
        if hasattr(main_window, 'server_section'):
            server_info = main_window.server_section.current_server_info
            if server_info:
                server_manager = DellServerManager(
                    ip=server_info['IP'],
                    port=server_info['PORT'],
                    auth=(server_info['USERNAME'], server_info['PASSWORD'])
                )
                progress_dialog.show()
                
                # 데이터 로드
                bios_info = server_manager.fetch_bios_info()
                idrac_info = server_manager.fetch_idrac_info()
                idrac_pwr_info = server_manager.fetch_idrac_pwr_info()
                nic_data = server_manager.fetch_network_adapters_info()

                # 섹션별 설정 딕셔너리 정의
                system_info_settings = {
                    'System Model Name': 'SystemModelName',
                    '제조사': 'SystemManufacturer',
                    'BIOS 버전': 'SystemBiosVersion',
                    'System Service Tag': 'SystemServiceTag'
                }
                
                processor_settings = {
                    'Logical Processor': 'LogicalProc',
                    'Virtualization Technology': 'ProcVirtualization',
                    'Sub NUMA Cluster': 'SubNumaCluster',
                    'x2APIC Mode': 'ProcX2Apic'
                }
                
                boot_settings = {
                    'Boot Mode': 'BootMode'
                }
                
                network_settings = {
                    'PXE Device1': 'PxeDev1EnDis',
                    'NIC ID 1': 'PxeDev1Interface',
                    'PXE Device2': 'PxeDev2EnDis',
                    'NIC ID 2': 'PxeDev2Interface',
                    'PXE Device3': 'PxeDev3EnDis',
                    'NIC ID 3': 'PxeDev3Interface',
                    'PXE Device4': 'PxeDev4EnDis',
                    'NIC ID 4': 'PxeDev4Interface'
                }
                
                integrated_devices = {
                    'SR-IOV Global Enable': 'SriovGlobalEnable',
                    'OS Watchdog Timer': 'OsWatchdogTimer'
                }

                system_profile_settings = {
                    'System Profile': 'SysProfile',
                    'CPU Power Management': 'ProcPwrPerf',
                    'Memory Frequency': 'ProcCStates',
                    'C1E': 'ProcC1E',
                    'Turbo Boost': 'ProcTurboMode',
                    'Energy Efficiency Policy': 'EnergyPerformanceBias',
                    'Memory Patrol Scrub': 'MemPatrolScrub',   
                }

                # AMD CPU 전용 설정
                amd_specific_settings = {
                    'Determinism Slider': 'DeterminismSlider',
                    'Power Profile Select': 'PowerProfileSelect',
                    'PCIE Speed PMM Control': 'PCIESpeedPMMControl',
                    'EQ Bypass To Highest Rate': 'EQBypassToHighestRate',
                    'DF PState Frequency Optimizer': 'DFPstateFrequencyOptimizer',
                    'DF PState Latency Optimizer': 'DFPstateLatencyOptimizer',
                    'DF CState': 'DfCState',
                    'Host System Management Port': 'HSMPSupport',
                    'Boost FMax': 'BoostFMax',
                    'Algorithm Performance Boost Disable': 'ApbDis',
                }

                miscellaneous_settings = {
                    'F1/F2 Prompt On Error': 'ErrPrompt'
                }

                idrac_settings = {
                    'Mac Address': 'NIC.1.MACAddress',
                    'Enable IPv4': 'IPv4.1.Enable',
                    'Enable DHCP': 'IPv4.1.DHCPEnable',
                    'Static IP Address': 'IPv4Static.1.Address',
                    'Static Gateway': 'IPv4Static.1.Gateway',
                    'Static Subnet Mask': 'IPv4Static.1.Netmask',
                    'Enable IPMI Over LAN': 'IPMILan.1.Enable',
                    'Enable VLAN ID': 'NIC.1.VLanEnable'
                }

                power_settings = {
                    'Redundancy Policy': 'ServerPwr.1.PSRedPolicy',
                    'Enable Hot Spare': 'ServerPwr.1.PSRapidOn'
                }

                nic_configuration_settings = {
                    'NIC ID': 'Id',
                    '가상화 모드': 'VirtualizationMode',
                    '링크 속도': 'LnkSpeed',
                    '부팅 프로토콜': 'LegacyBootProto'
                }

                # 섹션별 트리 아이템 생성
                sections = [
                    ("System Information", bios_info, system_info_settings),
                    ("Processor Settings", bios_info, processor_settings),
                    ("Boot Settings", bios_info, boot_settings),
                    ("Network Settings", bios_info, network_settings),
                    ("Integrated Devices", bios_info, integrated_devices),
                    ("System Profile Settings", bios_info, system_profile_settings),
                    ("Miscellaneous Settings", bios_info, miscellaneous_settings),
                    ("iDRAC Settings", idrac_info, idrac_settings),
                    ("Power Configuration", idrac_pwr_info, power_settings),
                    ("NIC Configuration", nic_data, nic_configuration_settings)
                ]

                for section_name, info_source, settings_dict in sections:
                    icon_base64 = get_icon_base64(section_name)
                    section_item = QTreeWidgetItem(tree_widget)
                    section_item.setText(0, section_name)
                    section_item.setIcon(0, QIcon(QPixmap.fromImage(QImage.fromData(base64.b64decode(icon_base64)))))
                    
                    if section_name == "NIC Configuration" and info_source and 'NetworkAdapters' in info_source:
                        for adapter in info_source['NetworkAdapters']:
                            for func in adapter.get('NetworkDeviceFunctions', []):
                                if func_id := func.get('Id'):
                                    virt_info = server_manager.fetch_network_virtualization_info(
                                        adapter.get('Id'), func_id)
                                    if virt_info and 'Attributes' in virt_info:
                                        attrs = virt_info['Attributes']
                                        
                                        # NIC 포트 아이템 생성
                                        port_item = QTreeWidgetItem(section_item)
                                        port_item.setText(0, func_id)
                                        
                                        # 가상화 모드
                                        virt_mode_item = QTreeWidgetItem(port_item)
                                        virt_mode_item.setText(1, "VirtualizationMode")
                                        virt_mode_item.setText(2, attrs.get('VirtualizationMode', 'N/A'))
                                        virt_mode_item.setToolTip(1, get_tooltip('VirtualizationMode'))
                                        
                                        # 링크 속도
                                        speed_item = QTreeWidgetItem(port_item)
                                        speed_item.setText(1, "LnkSpeed")
                                        speed_item.setText(2, attrs.get('LnkSpeed', 'N/A'))
                                        speed_item.setToolTip(1, get_tooltip('LnkSpeed'))
                                        
                                        # 부팅 프로토콜
                                        boot_item = QTreeWidgetItem(port_item)
                                        boot_item.setText(1, "LegacyBootProto")
                                        boot_item.setText(2, attrs.get('LegacyBootProto', 'N/A'))
                                        boot_item.setToolTip(1, get_tooltip('LegacyBootProto'))
                    
                    elif info_source and 'Attributes' in info_source:
                        # CPU 종류 확인
                        cpu_brand = info_source['Attributes'].get('Proc1Brand', '')
                        is_amd_cpu = 'AMD' in cpu_brand

                        # AMD CPU이고 System Profile Settings 섹션인 경우에만 추가 설정 병합
                        if is_amd_cpu and section_name == "System Profile Settings":
                            settings_dict.update(amd_specific_settings)

                        for display_name, attr_name in settings_dict.items():
                            value = info_source['Attributes'].get(attr_name, 'N/A')
                            
                            child_item = QTreeWidgetItem(section_item)
                            child_item.setText(0, display_name)
                            child_item.setText(1, attr_name)
                            child_item.setText(2, str(value))
                            
                            # 툴팁 추가
                            tooltip_text = get_tooltip(attr_name)
                            child_item.setToolTip(0, tooltip_text)
                            child_item.setToolTip(1, tooltip_text)
                            child_item.setToolTip(2, tooltip_text)
                            
                            # Enabled/Disabled 값에 따른 색상 설정
                            if value == 'Enabled':
                                child_item.setForeground(2, QColor('green'))
                            elif value == 'Disabled':
                                child_item.setForeground(2, QColor('red'))

                tree_widget.collapseAll()
                progress_dialog.setValue(100)
                
                def show_status_dialog():
                    progress_dialog.close()
                    status_dialog.exec()
                
                QTimer.singleShot(500, show_status_dialog)
                
    except Exception as e:
        progress_dialog.close()
        logger.error(f"시스템 정보 조회/표시 실패: {str(e)}")
        logger.exception(e)
        error_dialog = ErrorDialog(
            "시스템 정보 조회 오류",
            "시스템 정보를 조회하는 중 오류가 발생했습니다.",
            str(e),
            parent
        )
        error_dialog.exec()

def get_theme_color(parent):
    """현재 테마의 텍스트 색상 반환"""
    palette = parent.palette()
    return palette.text().color().name()

# NIC 정렬을 위한 순서 정의
def get_nic_order(component_id):
    """기본 타입 순서"""
    if 'NIC.Embedded.' in component_id:
        order = 0
    elif 'NIC.Integrated.' in component_id:
        order = 1
    elif 'NIC.Slot.' in component_id:
        order = 2
    else:
        return (3, 0, 0)  # 기타 항목
    
    """숫자 부분 추출 및 정렬을 위한 처리"""
    try:
        # NIC.Slot.1-2-1 또는 NIC.Embedded.1-1-1 등에서 숫자 추출
        parts = component_id.split('.')[-1].split('-')
        primary = int(parts[0])
        secondary = int(parts[1]) if len(parts) > 1 else 0
        tertiary = int(parts[2]) if len(parts) > 2 else 0
    except (ValueError, IndexError):
        primary = 0
        secondary = 0
        tertiary = 0
        
    return (order, primary, secondary, tertiary)

def get_icon_base64(section_name):
    """섹션 이름에 따라 적절한 아이콘의 Base64 인코딩 문자열 반환"""
    icon_map = {
        "System Information": "system_icon.png",
        "Processor Settings": "cpu_icon.png",
        "Boot Settings": "boot_icon.png",
        "Network Settings": "network_icon.png",
        "Integrated Devices": "device_icon.png",
        "System Profile Settings": "profile_icon.png",
        "Miscellaneous Settings": "misc_icon.png",
        "iDRAC Settings": "idrac_icon.png",
        "Power Configuration": "power_icon.png",
        "NIC Configuration": "nic_icon.png"
    }
    
    # 현재 스크립트의 디렉토리를 기준으로 상대 경로 설정
    current_dir = Path(os.path.dirname(os.path.abspath(__file__)))
    icon_dir = current_dir.parent.parent / "icon"
    icon_path = icon_dir / icon_map.get(section_name, "default_icon.png")
    
    try:
        with open(icon_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode("utf-8")
    except FileNotFoundError:
        print(f"아이콘 파일을 찾을 수 없습니다: {icon_path}")
        return ""  # 아이콘을 찾지 못한 경우 빈 문자열 반환

def get_section_settings(section_name, info_source):
    """섹션 이름에 따라 해당 섹션의 설정 딕셔너리 반환"""
    # 이미 정의된 섹션별 설정 딕셔너리 사용
    return globals().get(f"{section_name.lower().replace(' ', '_')}_settings", {})

def get_attribute_value(info_source, attr_name):
    """정보 소스에서 특정 속성의 값을 가져옴"""
    if isinstance(info_source, dict) and 'Attributes' in info_source:
        return info_source['Attributes'].get(attr_name, 'N/A')
    return 'N/A'

def get_value_style(value):
    """값에 따라 적절한 스타일 반환"""
    if value == 'Enabled':
        return "color: green;"
    elif value == 'Disabled':
        return "color: red;"
    return ""

def get_tooltip(attr_name):
    """속성 이름에 따른 툴팁 텍스트 반환"""
    tooltips = {
        # System Information
        "SystemModelName": "서버 모델 이름",
        "SystemManufacturer": "서버 제조사 정보",
        "SystemBiosVersion": "현재 설치된 BIOS 버전",
        "SystemServiceTag": "서버의 고유 식별 번호",
        
        # Processor Settings
        "LogicalProc": "논리 프로세서 활성화 여부",
        "ProcVirtualization": "프로세서 가상화 기술 활성화 여부",
        "SubNumaCluster": "NUMA 노드 내의 프로세서 코어 그룹 설정",
        "ProcX2Apic": "프로세서의 x2APIC 모드 활성화 여부",
        
        # Boot Settings
        "BootMode": "시스템 부팅 모드 (UEFI 또는 BIOS)",
        
        # Network Settings
        "PxeDev1EnDis": "첫 번째 PXE 장치 활성화/비활성화",
        "PxeDev1Interface": "첫 번째 PXE 장치의 네트워크 인터페이스",
        "PxeDev2EnDis": "두 번째 PXE 장치 활성화/비활성화",
        "PxeDev2Interface": "두 번째 PXE 장치의 네트워크 인터페이스",
        "PxeDev3EnDis": "세 번째 PXE 장치 활성화/비활성화",
        "PxeDev3Interface": "세 번째 PXE 장치의 네트워크 인터페이스",
        "PxeDev4EnDis": "네 번째 PXE 장치 활성화/비활성화",
        "PxeDev4Interface": "네 번째 PXE 장치의 네트워크 인터페이스",
        
        # Integrated Devices
        "SriovGlobalEnable": "SR-IOV 전역 활성화 여부",
        "OsWatchdogTimer": "운영 체제 감시 타이머 활성화 여부",
        
        # System Profile Settings
        "SysProfile": "시스템 프로필 설정",
        "ProcPwrPerf": "CPU 전원 관리 설정",
        "ProcCStates": "프로세서 C-States 설정",
        "ProcC1E": "프로세서 C1E 상태 활성화 여부",
        "ProcTurboMode": "프로세서 터보 부스트 모드 설정",
        "EnergyPerformanceBias": "에너지 효율성 정책 설정",
        "MemPatrolScrub": "메모리 패트롤 스크럽 기능 설정",
        
        # Miscellaneous Settings
        "ErrPrompt": "오류 발생 시 F1/F2 프롬프트 표시 여부",
        
        # iDRAC Settings
        "NIC.1.MACAddress": "iDRAC 네트워크 인터페이스의 MAC 주소",
        "IPv4.1.Enable": "IPv4 프로토콜 활성화 여부",
        "IPv4.1.DHCPEnable": "DHCP 사용 여부",
        "IPv4Static.1.Address": "고정 IP 주소 설정",
        "IPv4Static.1.Gateway": "고정 게이트웨이 주소",
        "IPv4Static.1.Netmask": "고정 서브넷 마스크",
        "IPMILan.1.Enable": "IPMI over LAN 활성화 여부",
        "NIC.1.VLanEnable": "VLAN 기능 활성화 여부",
        
        # Power Configuration
        "ServerPwr.1.PSRedPolicy": "전원 공급 장치 중복성 정책",
        "ServerPwr.1.PSRapidOn": "전원 공급 장치 핫 스페어 기능 활성화 여부",
        
        # NIC Configuration
        "VirtualizationMode": "NIC 가상화 모드 설정",
        "LnkSpeed": "NIC 링크 속도",
        "LegacyBootProto": "NIC 부팅 프로토콜"
    }
    return tooltips.get(attr_name, "설정에 대한 추가 정보")

def show_firmware_info(parent):
    """펌웨어 정보 조회"""
    logger.debug("펌웨어 정보 조회 시도")
    
    main_window = parent.window()
    if not hasattr(main_window, 'server_section'):
        logger.warning("서버 섹션을 찾을 수 없음")
        error_dialog = ErrorDialog(
            "서버 연결 오류",
            "서버가 연결되어 있지 않습니다.",
            "서버를 먼저 연결한 후 다시 시도해주세요.",
            parent
        )
        error_dialog.exec()
        return
    
    server_info = main_window.server_section.current_server_info
    if not server_info:
        logger.warning("서버 정보가 없음")
        error_dialog = ErrorDialog(
            "서버 연결 오류",
            "서버 정보를 찾을 수 없습니다.",
            "서버를 선택한 후 다시 시도해주세요.",
            parent
        )
        error_dialog.exec()
        return
    
    progress_dialog = QProgressDialog("펌웨어 정보 로딩 중...", None, 0, 100, parent)
    progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
    progress_dialog.setWindowTitle("데이터 로드")
    progress_dialog.setAutoClose(True)
    progress_dialog.setMinimumDuration(0)

    main_window = parent.window()
    if hasattr(main_window, 'server_section'):
        try:
            server_info = main_window.server_section.current_server_info
            if server_info:
                server_manager = DellServerManager(
                    ip=server_info['IP'],
                    port=server_info['PORT'],
                    auth=(server_info['USERNAME'], server_info['PASSWORD'])
                )
                
                progress_dialog.show()
                progress_dialog.setValue(30)
                
                firmware_data = server_manager.fetch_firmware_inventory()
                
                if firmware_data:
                    status_dialog = QDialog(parent)
                    status_dialog.setWindowTitle("펌웨어 정보")
                    status_dialog.resize(1000, 600)
                    layout = QVBoxLayout()

                    # 펌웨어 그룹 초기화
                    firmware_groups = {
                        'BIOS': [],
                        'iDRAC': [],
                        'RAID': [],
                        'NIC': [],
                        'Others': []
                    }

                    # 펌웨어 데이터를 그룹별로 분류
                    total_components = len(firmware_data.get('Members', []))
                    for idx, member in enumerate(firmware_data.get('Members', [])):
                        member_uri = member.get('@odata.id')
                        if member_uri:
                            component_id = member_uri.split('/')[-1]
                            component_info = server_manager.fetch_firmware_component(component_id)
                            
                            if 'BIOS' in component_id:
                                firmware_groups['BIOS'].append(component_info)
                            elif 'iDRAC' in component_id:
                                firmware_groups['iDRAC'].append(component_info)
                            elif 'PERC' in component_info.get('Name', ''):
                                firmware_groups['RAID'].append(component_info)
                            elif 'NIC' in component_id:
                                firmware_groups['NIC'].append(component_info)
                            else:
                                firmware_groups['Others'].append(component_info)
                            
                            progress_dialog.setValue(50 + (40 * idx // total_components))

                    # 테이블 위젯 생성
                    table_widget = QTableWidget()
                    table_widget.setColumnCount(6)
                    table_widget.setHorizontalHeaderLabels(["구성 요소", "버전", "상태", "날짜", "재시작 필요", "비고"])
                    
                    # 컬럼 너비 설정
                    table_widget.setColumnWidth(0, 250)  # 구성 요소
                    table_widget.setColumnWidth(1, 150)  # 버전
                    table_widget.setColumnWidth(2, 100)  # 상태
                    table_widget.setColumnWidth(3, 150)  # 날짜
                    table_widget.setColumnWidth(4, 100)  # 재시작 필요
                    table_widget.setColumnWidth(5, 200)  # 비고
                    
                    # 테이블 스타일 설정
                    table_widget.setAlternatingRowColors(True)
                    table_widget.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
                    table_widget.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

                    row = 0
                    restart_required = False
                    for group_name, components in firmware_groups.items():
                        if components:
                            # 그룹 헤더 추가
                            table_widget.insertRow(row)
                            header_item = QTableWidgetItem(group_name)
                            header_item.setBackground(QColor("#E3F2FD"))
                            for col in range(6):
                                table_widget.setItem(row, col, QTableWidgetItem(""))
                                table_widget.item(row, col).setBackground(QColor("#E3F2FD"))
                            table_widget.setItem(row, 0, header_item)
                            row += 1
                            
                            # 현재 버전과 이전 버전 컴포넌트 분리
                            current_components = []
                            previous_components = []
                            
                            for component in components:
                                component_id = component.get('Id', '')
                                if 'Installed' in component_id:
                                    current_components.append(component)
                                elif 'Previous' in component_id:
                                    previous_components.append(component)

                            # 현재 버전 컴포넌트 추가
                            for component in current_components:
                                table_widget.insertRow(row)
                                
                                # 구성 요소
                                name_item = QTableWidgetItem(component.get('Name', 'Unknown'))
                                table_widget.setItem(row, 0, name_item)
                                
                                # 버전
                                version_item = QTableWidgetItem(component.get('Version', 'Unknown'))
                                table_widget.setItem(row, 1, version_item)
                                
                                # 상태
                                status = component.get('Status', {}).get('Health', 'Unknown')
                                status_item = QTableWidgetItem(status)
                                if status == 'OK':
                                    status_item.setForeground(QColor("#2E7D32"))
                                elif status == 'Warning':
                                    status_item.setForeground(QColor("#F57F17"))
                                elif status == 'Critical':
                                    status_item.setForeground(QColor("#B71C1C"))
                                table_widget.setItem(row, 2, status_item)
                                
                                # 설치 날짜
                                install_date = component.get('Oem', {}).get('Dell', {}).get(
                                    'DellSoftwareInventory', {}).get('InstallationDate', 'Unknown')
                                if install_date and install_date != 'Unknown':
                                    date_parts = install_date.split('T')
                                    if len(date_parts) == 2:
                                        install_date = f"{date_parts[0]} {date_parts[1][:5]}"
                                date_item = QTableWidgetItem(install_date)
                                table_widget.setItem(row, 3, date_item)
                                
                                # 재시작 필요 여부
                                needs_restart = component.get('RebootRequired', False)
                                restart_required = restart_required or needs_restart
                                restart_item = QTableWidgetItem('예' if needs_restart else '아니오')
                                table_widget.setItem(row, 4, restart_item)
                                
                                # 현재 설치됨 표시
                                note_item = QTableWidgetItem("현재 설치됨")
                                note_item.setForeground(QColor("#2E7D32"))
                                table_widget.setItem(row, 5, note_item)
                                
                                row += 1

                            # 이전 버전 컴포넌트 추가
                            for component in previous_components:
                                table_widget.insertRow(row)
                                
                                # 구성 요소 (회색으로 표시)
                                name_item = QTableWidgetItem(component.get('Name', 'Unknown'))
                                name_item.setForeground(QColor("#666666"))
                                table_widget.setItem(row, 0, name_item)
                                
                                # 버전
                                version_item = QTableWidgetItem(component.get('Version', 'Unknown'))
                                version_item.setForeground(QColor("#666666"))
                                table_widget.setItem(row, 1, version_item)
                                
                                # 상태
                                if component.get('Status'):
                                    status = component.get('Status', {}).get('Health', 'Unknown')
                                    status_item = QTableWidgetItem(status)
                                    table_widget.setItem(row, 2, status_item)
                                
                                # 마지막 사용 날짜
                                last_date = component.get('Oem', {}).get('Dell', {}).get(
                                    'DellSoftwareInventory', {}).get('LastInstallationDate', 'Unknown')
                                if last_date and last_date != 'Unknown':
                                    date_parts = last_date.split('T')
                                    if len(date_parts) == 2:
                                        last_date = f"{date_parts[0]} {date_parts[1][:5]}"
                                date_item = QTableWidgetItem(last_date)
                                date_item.setForeground(QColor("#666666"))
                                table_widget.setItem(row, 3, date_item)
                                
                                # 롤백 가능 표시
                                note_item = QTableWidgetItem("롤백 가능")
                                note_item.setForeground(QColor("#1976D2"))
                                table_widget.setItem(row, 5, note_item)
                                
                                row += 1

                            # 그룹 사이에 빈 줄 추가
                            table_widget.insertRow(row)
                            for col in range(6):
                                table_widget.setItem(row, col, QTableWidgetItem(""))
                            row += 1
                    
                    def show_update_dialog():
                        file_dialog = QFileDialog()
                        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFiles)  # 다중 선택 모드
                        
                        # 마지막 디렉토리가 있으면 해당 위치에서 시작
                        if hasattr(status_dialog, 'last_firmware_directory') and status_dialog.last_firmware_directory and os.path.exists(status_dialog.last_firmware_directory):
                            file_dialog.setDirectory(status_dialog.last_firmware_directory)
                        
                        file_paths, _ = file_dialog.getOpenFileNames(
                            parent,
                            "펌웨어 이미지 선택",
                            "",
                            "펌웨어 이미지 (*.exe *.EXE *.BIN *.bin *.upm *.UPM *.pmc *.PMC)"
                        )
                        
                        if file_paths:
                            # 선택된 디렉토리 저장
                            status_dialog.last_firmware_directory = os.path.dirname(file_paths[0])
                            
                            # 선택된 파일 목록을 보여주는 확인 다이얼로그
                            files_text = "\n".join([f"- {os.path.basename(path)}" for path in file_paths])
                            confirm = QMessageBox.question(
                                parent,
                                "펌웨어 업데이트 확인",
                                f"다음 파일들로 펌웨어 업데이트를 진행하시겠습니까?\n\n{files_text}",
                                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                            )
                            if confirm == QMessageBox.StandardButton.Yes:
                                try:
                                    if len(file_paths) == 1:
                                        # 단일 파일 업데이트
                                        result = server_manager.update_firmware(file_paths[0])
                                    else:
                                        # 멀티파트 업데이트
                                        result = server_manager.multipart_firmware_update(file_paths)
                                    
                                    if result:
                                        QMessageBox.information(
                                            parent,
                                            "업데이트 시작",
                                            "펌웨어 업데이트가 시작되었습니다. 작업 큐에서 진행 상황을 확인하세요."
                                        )
                                except Exception as e:
                                    QMessageBox.critical(
                                        parent,
                                        "업데이트 오류",
                                        f"펌웨어 업데이트 중 오류가 발생했습니다: {str(e)}"
                                    )

                    def show_rollback_dialog():
                        # 선택된 행 가져오기
                        selected_rows = set(item.row() for item in table_widget.selectedItems())
                        if not selected_rows:
                            QMessageBox.warning(
                                status_dialog,
                                "경고",
                                "롤백할 펌웨어를 선택해주세요.",
                                QMessageBox.StandardButton.Ok
                            )
                            return

                        # 선택된 각 행에 대해 처리
                        for row in selected_rows:
                            # 구성 요소 이름과 버전 가져오기
                            component_name = table_widget.item(row, 0).text()
                            component_version = table_widget.item(row, 1).text()
                            note = table_widget.item(row, 5).text()
                            
                            # 현재 설치된 버전은 롤백 불가
                            if note == "현재 설치됨":
                                QMessageBox.warning(
                                    status_dialog,
                                    "경고",
                                    f"{component_name}은(는) 현재 설치된 버전이므로 롤백할 수 없습니다.",
                                    QMessageBox.StandardButton.Ok
                                )
                                continue

                            # 롤백 확인 메시지
                            reply = QMessageBox.question(
                                status_dialog,
                                "펌웨어 롤백",
                                f"선택한 펌웨어를 롤백하시겠습니까?\n\n"
                                f"구성 요소: {component_name}\n"
                                f"버전: {component_version}",
                                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                QMessageBox.StandardButton.No
                            )

                            if reply == QMessageBox.StandardButton.Yes:
                                # 여기에 실제 롤백 로직 구현
                                print(f"Rolling back {component_name} to version {component_version}")
                                QMessageBox.information(
                                    status_dialog,
                                    "알림",
                                    "롤백이 시작되었습니다.\n"
                                    "작업 관리 탭에서 진행 상황을 확인할 수 있습니다.",
                                    QMessageBox.StandardButton.Ok
                                )

                    def show_queue_dialog():  # parent 매개변수 제거
                        """작업 큐 관리 대화상자를 표시합니다."""
                        dialog = QDialog(status_dialog)
                        dialog.setWindowTitle("작업 관리")
                        dialog.resize(900, 500)
                        
                        # 메인 레이아웃
                        layout = QVBoxLayout()
                        
                        # 상단 필터 영역
                        filter_layout = QHBoxLayout()
                        
                        # 상태 필터
                        status_label = QLabel("상태:")
                        status_combo = QComboBox()
                        status_combo.addItems(["전체", "대기 중", "진행 중", "완료", "실패"])
                        filter_layout.addWidget(status_label)
                        filter_layout.addWidget(status_combo)
                        
                        # 작업 종류 필터
                        type_label = QLabel("작업 종류:")
                        type_combo = QComboBox()
                        type_combo.addItems(["전체", "펌웨어 업데이트", "펌웨어 롤백", "재시작"])
                        filter_layout.addWidget(type_label)
                        filter_layout.addWidget(type_combo)
                        
                        # 필터 레이아웃을 메인 레이아웃에 추가
                        layout.addLayout(filter_layout)
                        
                        # 작업 목록 테이블
                        table = QTableWidget()
                        table.setColumnCount(7)
                        table.setHorizontalHeaderLabels([
                            "작업 ID", "작업 종류", "구성 요소", "상태", 
                            "진행률", "시작 시각", "예정된 재시작"
                        ])
                        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
                        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

                        # 실제 작업 데이터 가져오기
                        try:
                            server_manager = DellServerManager(
                                ip=server_info['IP'],
                                port=server_info['PORT'],
                                auth=(server_info['USERNAME'], server_info['PASSWORD'])
                            )
                            queue_data = server_manager.get_firmware_queue()
                            
                            if queue_data and 'Members' in queue_data:
                                jobs = queue_data['Members']
                                table.setRowCount(len(jobs))
                                for row, job in enumerate(jobs):
                                    # 작업 ID
                                    table.setItem(row, 0, QTableWidgetItem(job.get('Id', '')))
                                    
                                    # 작업 종류
                                    job_type = ''
                                    job_name = job.get('Name', '').lower()
                                    if 'update' in job_name:
                                        job_type = '펌웨어 업데이트'
                                    elif 'rollback' in job_name:
                                        job_type = '펌웨어 롤백'
                                    elif 'restart' in job_name:
                                        job_type = '재시작'
                                    table.setItem(row, 1, QTableWidgetItem(job_type))
                                    
                                    # 구성 요소
                                    table.setItem(row, 2, QTableWidgetItem(job.get('Component', '')))
                                    
                                    # 상태
                                    status = job.get('JobState', '')
                                    status_item = QTableWidgetItem(status)
                                    if status == '완료':
                                        status_item.setForeground(QColor("#2E7D32"))
                                    elif status == '진행 중':
                                        status_item.setForeground(QColor("#1976D2"))
                                    elif status == '실패':
                                        status_item.setForeground(QColor("#B71C1C"))
                                    table.setItem(row, 3, status_item)
                                    
                                    # 진행률
                                    progress = job.get('PercentComplete', '0')
                                    table.setItem(row, 4, QTableWidgetItem(f"{progress}%"))
                                    
                                    # 시작 시각
                                    start_time = job.get('StartTime', '')
                                    table.setItem(row, 5, QTableWidgetItem(start_time))
                                    
                                    # 예정된 재시작
                                    reboot_time = job.get('RebootTime', '')
                                    table.setItem(row, 6, QTableWidgetItem(reboot_time))
                            else:
                                QMessageBox.information(
                                    dialog,
                                    "알림",
                                    "현재 진행 중인 작업이 없습니다."
                                )
                                
                        except Exception as e:
                            logger.error(f"작업 목록 조회 실패: {str(e)}")
                            ErrorDialog(
                                "작업 목록 조회 실패",
                                "작업 목록을 가져오는데 실패했습니다.",
                                str(e),
                                parent
                            ).exec()
                        
                        # 하단 버튼
                        button_layout = QHBoxLayout()
                        refresh_btn = QPushButton("새로고침")
                        refresh_btn.setFixedWidth(150)
                        cancel_job_btn = QPushButton("작업 취소")
                        cancel_job_btn.setFixedWidth(150)
                        
                        button_layout.addWidget(refresh_btn)
                        button_layout.addWidget(cancel_job_btn)
                        layout.addLayout(button_layout)

                        def refresh_job_list():
                            """작업 목록을 새로고침합니다."""
                            try:
                                queue_data = server_manager.get_firmware_queue()
                                table.clearContents()
                                if queue_data and 'Members' in queue_data:
                                    jobs = queue_data['Members']
                                    table.setRowCount(len(jobs))
                                    # ... (위의 작업 목록 표시 코드와 동일)
                                else:
                                    table.setRowCount(0)
                                    QMessageBox.information(
                                        dialog,
                                        "알림",
                                        "현재 진행 중인 작업이 없습니다."
                                    )
                            except Exception as e:
                                logger.error(f"작업 목록 새로고침 실패: {str(e)}")
                                ErrorDialog(
                                    "새로고침 실패",
                                    "작업 목록을 새로고침하는데 실패했습니다.",
                                    str(e),
                                    parent
                                ).exec()

                        def cancel_selected_job():
                            """선택된 작업을 취소합니다."""
                            selected_rows = table.selectedItems()
                            if not selected_rows:
                                QMessageBox.warning(
                                    dialog,
                                    "경고",
                                    "취소할 작업을 선택해주세요."
                                )
                                return
                            
                            job_id = table.item(table.currentRow(), 0).text()
                            status = table.item(table.currentRow(), 3).text()
                            
                            if status == '완료' or status == '실패':
                                QMessageBox.warning(
                                    dialog,
                                    "경고",
                                    "이미 완료되거나 실패한 작업은 취소할 수 없습니다."
                                )
                                return
                            
                            reply = QMessageBox.question(
                                dialog,
                                "작업 취소 확인",
                                f"선택한 작업(ID: {job_id})을 취소하시겠습니까?",
                                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                QMessageBox.StandardButton.No
                            )

                            if reply == QMessageBox.StandardButton.Yes:
                                try:
                                    server_manager.cancel_firmware_job(job_id)
                                    QMessageBox.information(
                                        dialog,
                                        "작업 취소 완료",
                                        "작업이 취소되었습니다."
                                    )
                                    refresh_job_list()
                                except Exception as e:
                                    logger.error(f"작업 취소 실패: {str(e)}")
                                    ErrorDialog(
                                        "작업 취소 실패",
                                        "작업을 취소하는데 실패했습니다.",
                                        str(e),
                                        parent
                                    ).exec()

                        # 버튼 연결
                        refresh_btn.clicked.connect(lambda _: refresh_job_list())
                        cancel_job_btn.clicked.connect(lambda _: cancel_selected_job())
                        
                        dialog.setLayout(layout)
                        dialog.exec()

                    # 테이블 위젯 추가 후
                    layout.addWidget(table_widget)
                    
                    # 하단 버튼들
                    button_layout = QHBoxLayout()

                    # 펌웨어 업데이트 버튼
                    update_button = QPushButton("fw update")
                    update_button.setFixedWidth(150)
                    button_layout.addWidget(update_button)

                    # 펌웨어 로트백 버튼
                    rollback_button = QPushButton("fw rollback")
                    rollback_button.setFixedWidth(150)
                    button_layout.addWidget(rollback_button)

                    # 작업 관리 버튼
                    queue_button = QPushButton("fw queue")
                    queue_button.clicked.connect(show_queue_dialog)
                    queue_button.setFixedWidth(150)
                    button_layout.addWidget(queue_button)

                    # 재시작 관련 버튼들을 추가
                    if restart_required:
                        restart_label = QLabel("일부 변경사항은 시스템 재시작이 필요합니다.")
                        restart_label.setStyleSheet("color: red;")
                        layout.addWidget(restart_label)
                        
                        schedule_restart_btn = QPushButton("재시작 예약")
                        schedule_restart_btn.clicked.connect(show_restart_scheduler)
                        schedule_restart_btn.setFixedWidth(150)
                        
                        immediate_restart_btn = QPushButton("즉시 재시작")
                        immediate_restart_btn.clicked.connect(confirm_immediate_restart)
                        immediate_restart_btn.setFixedWidth(150)
                        
                        button_layout.addWidget(schedule_restart_btn)
                        button_layout.addWidget(immediate_restart_btn)
                    
                    layout.addLayout(button_layout)

                    # 버튼 연결
                    update_button.clicked.connect(show_update_dialog)
                    rollback_button.clicked.connect(show_rollback_dialog)

                    status_dialog.setLayout(layout)
                    status_dialog.exec()
            
            progress_dialog.close()
            
        except Exception as e:
            progress_dialog.close()
            error_dialog = ErrorDialog(
                "오류 발생",
                "펌웨어 정보를 불러오는 중 오류가 발생했습니다.",
                str(e),
                parent
            )
            error_dialog.exec()
            logger.error(f"펌웨어 정보 조회 중 오류 발생: {str(e)}")

def sort_drives(drive_info):
    import re  # 명시적 import 추가
    def sort_key(drive):
        # Disk.Bay.숫자:Enclosure... 형식에서 숫자만 추출하여 정렬
        match = re.search(r"Disk\.Bay\.(\d+)", drive.get('Id', ''))
        if match:
            return int(match.group(1))
        return float('inf')

    return sorted(drive_info, key=sort_key)

def show_restart_scheduler(parent):
    """재시작 일정 예약 다이얼로그를 표시합니다."""
    dialog = QDialog(parent)
    dialog.setWindowTitle("시스템 재시작 예약")
    layout = QVBoxLayout()
    
    # 날짜/시간 선택 위젯
    date_time_edit = QDateTimeEdit()
    date_time_edit.setDateTime(QDateTime.currentDateTime().addSecs(3600))  # 기본값: 1시간 후
    date_time_edit.setMinimumDateTime(QDateTime.currentDateTime())
    date_time_edit.setCalendarPopup(True)
    layout.addWidget(QLabel("재시작 시간 선택:"))
    layout.addWidget(date_time_edit)
    
    # 버튼
    button_box = QDialogButtonBox(
        QDialogButtonBox.StandardButton.Ok | 
        QDialogButtonBox.StandardButton.Cancel
    )
    button_box.accepted.connect(lambda: schedule_restart(parent, date_time_edit.dateTime()))
    button_box.accepted.connect(dialog.accept)
    button_box.rejected.connect(dialog.reject)
    layout.addWidget(button_box)
    
    dialog.setLayout(layout)
    dialog.exec()

def confirm_immediate_restart(parent):
    """즉시 재시작 확인 다이얼로그를 표시합니다."""
    msg_box = QMessageBox(parent)
    msg_box.setIcon(QMessageBox.Icon.Warning)
    msg_box.setWindowTitle("시스템 재시작 확인")
    msg_box.setText("시스템을 즉시 재시작하시겠습니까?")
    msg_box.setInformativeText("모든 작업이 중단되며, 재시작이 완료될 때까지 시스템에 접근할 수 없습니다.")
    msg_box.setStandardButtons(
        QMessageBox.StandardButton.Yes | 
        QMessageBox.StandardButton.No
    )
    msg_box.setDefaultButton(QMessageBox.StandardButton.No)
    
    if msg_box.exec() == QMessageBox.StandardButton.Yes:
        perform_restart(parent)

def schedule_restart(parent, restart_time):
    """시스템 재시작을 예약합니다."""
    main_window = parent.window()
    if hasattr(main_window, 'server_section'):
        server_info = main_window.server_section.current_server_info
        if server_info:
            try:
                server_manager = DellServerManager(
                    ip=server_info['IP'],
                    port=server_info['PORT'],
                    auth=(server_info['USERNAME'], server_info['PASSWORD'])
                )
                # 재시작 예약 API 호출
                server_manager.schedule_system_restart(restart_time)
                QMessageBox.information(
                    parent,
                    "재시작 예약 완료",
                    f"시스템 재시작이 {restart_time.toString('yyyy-MM-dd hh:mm')}에 예약되었습니다."
                )
            except Exception as e:
                logger.error(f"재시작 예약 실패: {str(e)}")
                ErrorDialog(
                    "재시작 예약 실패",
                    "시스템 재시작 예약에 실패했습니다.",
                    str(e),
                    parent
                ).exec()

def perform_restart(parent):
    """시스템을 즉시 재시작합니다."""
    main_window = parent.window()
    if hasattr(main_window, 'server_section'):
        server_info = main_window.server_section.current_server_info
        if server_info:
            try:
                server_manager = DellServerManager(
                    ip=server_info['IP'],
                    port=server_info['PORT'],
                    auth=(server_info['USERNAME'], server_info['PASSWORD'])
                )
                # 즉시 재시작 API 호출
                server_manager.restart_system()
                QMessageBox.information(
                    parent,
                    "재시작 시작",
                    "시스템 재시작이 시작되었습니다. 잠시 후 다시 연결해주세요."
                )
            except Exception as e:
                logger.error(f"재시작 실패: {str(e)}")
                ErrorDialog(
                    "재시작 실패",
                    "시스템 재시작에 실패했습니다.",
                    str(e),
                    parent
                ).exec()

def show_log_popup(parent, log_type):
    logger.debug(f"{log_type.upper()} 로그 팝업창 열기 시도")
    
    main_window = parent.window()
    if not hasattr(main_window, 'server_section'):
        logger.warning("서버 섹션을 찾을 수 없음")
        error_dialog = ErrorDialog(
            "서버 연결 오류",
            "서버가 연결되어 있지 않습니다.",
            "서버를 먼저 연결한 후 다시 시도해주세요.",
            parent
        )
        error_dialog.exec()
        return
    
    server_info = main_window.server_section.current_server_info
    if not server_info:
        logger.warning("서버 정보가 없음")
        error_dialog = ErrorDialog(
            "서버 연결 오류",
            "서버 정보를 찾을 수 없습니다.",
            "서버를 선택한 후 다시 시도해주세요.",
            parent
        )
        error_dialog.exec()
        return

    try:
        server_manager = DellServerManager(
            ip=server_info['IP'],
            port=server_info['PORT'],
            auth=(server_info['USERNAME'], server_info['PASSWORD'])
        )

        dialog = QDialog(parent)
        dialog.setWindowTitle(f"{log_type.upper()} 로그")
        dialog.resize(1200, 800)
        main_layout = QVBoxLayout(dialog)

        # 탭 위젯 추가
        tab_widget = QTabWidget()
        main_layout.addWidget(tab_widget)

        # 로그 뷰어 탭
        log_viewer_tab = QWidget()
        log_viewer_layout = QVBoxLayout(log_viewer_tab)
        
        # 필터 컨트롤 추가
        filter_layout = QHBoxLayout()
        
        # 심각도 필터 초기화를 안전하게 수행
        severity_combo = QComboBox()
        severity_items = ['전체', 'Critical', 'Warning', 'OK']
        for item in severity_items:
            severity_combo.addItem(item)
        
        # 검색 입력
        search_input = QLineEdit()
        search_input.setPlaceholderText("메시지 내용으로 검색")
        
        filter_layout.addWidget(QLabel("심각도:"))
        filter_layout.addWidget(severity_combo)
        filter_layout.addWidget(QLabel("검색:"))
        filter_layout.addWidget(search_input)
        log_viewer_layout.addLayout(filter_layout)

        # 로그 목록 트리 위젯
        tree_widget = QTreeWidget(dialog)
        tree_widget.setHeaderLabels(["ID", "심각도", "생성 시간", "메시지"])
        
        # 컬럼 너비 최적화
        tree_widget.setColumnWidth(0, 100)   # ID
        tree_widget.setColumnWidth(1, 100)   # 심각도
        tree_widget.setColumnWidth(2, 150)   # 생성 시간
        tree_widget.setColumnWidth(3, 600)   # 메시지
        log_viewer_layout.addWidget(tree_widget)

        # 버튼 레이아웃
        button_layout = QHBoxLayout()
        refresh_button = QPushButton("새로고침")
        copy_button = QPushButton("로그 복사")
        excel_button = QPushButton("Excel 내보내기")
        
        # SEL 로그일 경우에만 클리어 버튼 추가
        if log_type == 'sel':
            clear_button = QPushButton("로그 클리어")
            clear_button.setIcon(QIcon("clear_icon.png"))
        
        button_layout.addWidget(refresh_button)
        button_layout.addWidget(copy_button)
        button_layout.addWidget(excel_button)
        if log_type == 'sel':
            button_layout.addWidget(clear_button)
        button_layout.addStretch()
        log_viewer_layout.addLayout(button_layout)

        # 로그 분석 탭
        log_analysis_tab = QWidget()
        log_analysis_layout = QVBoxLayout(log_analysis_tab)
        
        # 로그 레벨 통계 섹션
        log_level_stats_label = QLabel("로그 레벨 통계")
        log_level_stats_label.setStyleSheet("font-weight: bold; font-size: 16px;")
        log_analysis_layout.addWidget(log_level_stats_label)
        
        # 로그 레벨 차트를 위한 스크롤 영역
        log_level_scroll_area = QScrollArea()
        log_level_scroll_area.setWidgetResizable(True)
        log_level_chart_widget = QWidget()
        log_level_chart_layout = QVBoxLayout(log_level_chart_widget)
        log_level_scroll_area.setWidget(log_level_chart_widget)
        log_analysis_layout.addWidget(log_level_scroll_area)
        
        # 타임라인 통계 섹션
        timeline_stats_label = QLabel("시간대별 로그 통계")
        timeline_stats_label.setStyleSheet("font-weight: bold; font-size: 16px;")
        log_analysis_layout.addWidget(timeline_stats_label)
        
        # 타임라인 차트를 위한 스크롤 영역
        timeline_scroll_area = QScrollArea()
        timeline_scroll_area.setWidgetResizable(True)
        timeline_chart_widget = QWidget()
        timeline_chart_layout = QVBoxLayout(timeline_chart_widget)
        timeline_scroll_area.setWidget(timeline_chart_widget)
        log_analysis_layout.addWidget(timeline_scroll_area)

        # 로그 엔트리 저장 리스트
        log_entries = []

        def add_log_to_tree(log_entry):
            nonlocal log_entries
            item = QTreeWidgetItem(tree_widget)
            
            # ID 설정
            item.setText(0, log_entry.get('Id', 'N/A'))
            
            # 심각도 설정 및 색상 적용
            severity = log_entry.get('Severity', 'N/A')
            item.setText(1, severity)
            if severity == 'Critical':
                item.setForeground(1, QColor('red'))
            elif severity == 'Warning':
                item.setForeground(1, QColor('orange'))
            elif severity == 'OK':
                item.setForeground(1, QColor('green'))
            
            # 시간 형식 변환
            created_time = log_entry.get('Created', 'N/A')
            item.setText(2, format_time(created_time))
            
            # 메시지
            item.setText(3, log_entry.get('Message', 'N/A'))
            
            log_entries.append(log_entry)

        def calculate_log_statistics(entries):
            # 로그 엔트리가 없으면 빈 그래프 생성
            if not entries:
                # 로그 레벨 통계 그래프
                plt.close('all')
                fig, ax = plt.subplots(figsize=(8, 3))
                ax.text(0.5, 0.5, '로그 데이터 없음', 
                        horizontalalignment='center', 
                        verticalalignment='center')
                plt.title('로그 레벨 분포', fontsize=10)
                canvas = FigureCanvas(fig)
                
                # 기존 위젯 제거 및 새 캔버스 추가
                for i in reversed(range(log_level_chart_layout.count())): 
                    log_level_chart_layout.itemAt(i).widget().setParent(None)
                log_level_chart_layout.addWidget(canvas)
                
                # 타임라인 통계 그래프
                plt.close('all')
                fig, ax = plt.subplots(figsize=(8, 3))
                ax.text(0.5, 0.5, '로그 데이터 없음', 
                        horizontalalignment='center', 
                        verticalalignment='center')
                plt.title('시간대별 로그 분포', fontsize=10)
                canvas = FigureCanvas(fig)
                
                # 기존 위젯 제거 및 새 캔버스 추가
                for i in reversed(range(timeline_chart_layout.count())): 
                    timeline_chart_layout.itemAt(i).widget().setParent(None)
                timeline_chart_layout.addWidget(canvas)
                return
            
            # matplotlib 한글 폰트 설정
            get_system_matplotlib_font()
            
            # 로그 레벨 통계
            severity_counts = Counter(entry.get('Severity', 'N/A') for entry in entries)
            total_entries = len(entries)
            
            # 색상 매핑
            color_map = {
                'Critical': '#FF6384',   # 진한 빨간색
                'Warning': '#FFCE56',    # 노란색
                'OK': '#4BC0C0'          # 청록색
            }
            
            # 새 수평 막대 그래프 생성
            plt.close('all')  # 기존 플롯 닫기
            
            # 데이터 개수에 따라 동적으로 그래프 높이 조정
            graph_height = max(3, min(len(severity_counts) * 0.5, 6))
            if len(severity_counts) <= 2:
                graph_height = 2.5  # 로그 개수가 적을 때 더 작은 높이
            fig, ax = plt.subplots(figsize=(8, graph_height))
            
            # 데이터 준비
            levels = list(severity_counts.keys())
            counts = list(severity_counts.values())
            percentages = [(count / total_entries * 100) for count in counts]
            
            colors = [color_map.get(level, '#000000') for level in levels]
            
            # 수평 막대 그래프 생성
            bars = ax.barh(levels, percentages, color=colors, height=0.5)  # 막대 높이 더 작게
            
            # 각 막대에 로그 개수 표시
            for bar, count in zip(bars, counts):
                width = bar.get_width()
                ax.text(width, bar.get_y() + bar.get_height()/2, 
                        f'{count}', 
                        va='center', fontsize=8)
            
            plt.title('로그 레벨 분포', fontsize=10)
            plt.xlabel('비율 (%)', fontsize=9)
            plt.xticks(fontsize=8)
            plt.yticks(fontsize=8)
            plt.tight_layout()
            
            # 기존 위젯 제거 및 새 캔버스 추가
            for i in reversed(range(log_level_chart_layout.count())): 
                log_level_chart_layout.itemAt(i).widget().setParent(None)
            canvas = FigureCanvas(fig)
            log_level_chart_layout.addWidget(canvas)

            # 타임라인 통계
            # 시간대별 로그 분포
            timeline_counts = {}
            for entry in entries:
                try:
                    entry_time = datetime.fromisoformat(entry.get('Created', '').replace('Z', '+00:00'))
                    hour_key = entry_time.strftime("%Y/%m/%d %H시")
                    timeline_counts[hour_key] = timeline_counts.get(hour_key, 0) + 1
                except:
                    pass
            
            # 시간대 순서대로 정렬
            sorted_timeline = dict(sorted(timeline_counts.items(), key=lambda x: x[0]))
            
            # 새 막대 그래프 생성
            plt.close('all')  # 기존 플롯 닫기
            
            # 데이터 개수에 따라 동적으로 그래프 높이 조정
            graph_height = max(3, min(len(sorted_timeline) * 0.5, 6))
            if len(sorted_timeline) <= 2:
                graph_height = 2.5  # 로그 개수가 적을 때 더 작은 높이
            fig, ax = plt.subplots(figsize=(8, graph_height))
            
            # 데이터 준비
            time_periods = list(sorted_timeline.keys())
            counts = list(sorted_timeline.values())
            percentages = [(count / total_entries * 100) for count in counts]
            
            # 막대 그래프 생성
            bars = ax.barh(time_periods, counts, color='#4BC0C0', height=0.5)  # 막대 높이 더 작게
            
            # 각 막대에 로그 개수 표시
            for bar, count in zip(bars, counts):
                width = bar.get_width()
                ax.text(width, bar.get_y() + bar.get_height()/2, 
                        f'{count}', 
                        va='center', fontsize=8)
            
            plt.title('시간대별 로그 분포', fontsize=10)
            plt.xlabel('로그 수', fontsize=9)
            plt.xticks(fontsize=8)
            plt.yticks(fontsize=8)
            plt.tight_layout()
            
            # 기존 위젯 제거 및 새 캔버스 추가
            for i in reversed(range(timeline_chart_layout.count())): 
                timeline_chart_layout.itemAt(i).widget().setParent(None)
            canvas = FigureCanvas(fig)
            timeline_chart_layout.addWidget(canvas)

        def refresh_logs():
            # log_entries 초기화
            nonlocal log_entries
            log_entries = []
            
            tree_widget.clear()
            try:
                # 로그 타입에 따라 적절한 메서드 호출
                if log_type == 'sel':
                    log_data = server_manager.fetch_sel_entries()
                else:  # log_type == 'lc'
                    log_data = server_manager.fetch_lc_entries()
                
                # log_data가 None이면 빈 리스트로 처리
                entries = log_data.get('Members', []) if log_data else []
                
                # 필터링 적용
                filtered_entries = []
                severity_filter = severity_combo.currentText()
                search_text = search_input.text().lower()
                
                for entry in entries:
                    # 심각도 필터 적용
                    if severity_filter != '전체' and entry.get('Severity') != severity_filter:
                        continue
                    
                    # 검색어 필터 적용
                    if search_text and search_text not in entry.get('Message', '').lower():
                        continue
                    
                    filtered_entries.append(entry)
                    add_log_to_tree(entry)
                
                # 로그 통계 계산 및 표시
                calculate_log_statistics(filtered_entries)
                
            except Exception as e:
                QMessageBox.critical(dialog, "오류", f"로그 조회 실패: {str(e)}")
                # 빈 리스트로 통계 그래프 생성
                calculate_log_statistics([])
        
        def copy_logs_to_clipboard():
            if not log_entries:
                QMessageBox.warning(dialog, "경고", "복사할 로그가 없습니다.")
                return
            
            clipboard_text = "\n".join([
                f"ID: {entry.get('Id', 'N/A')} | "
                f"심각도: {entry.get('Severity', 'N/A')} | "
                f"시간: {format_time(entry.get('Created', 'N/A'))} | "
                f"메시지: {entry.get('Message', 'N/A')}"
                for entry in log_entries
            ])
            
            clipboard = QApplication.clipboard()
            clipboard.setText(clipboard_text)
            QMessageBox.information(dialog, "완료", "로그가 클립보드에 복사되었습니다.")

        def export_logs_to_xlsx():
            if not log_entries:
                QMessageBox.warning(dialog, "경고", "내보낼 로그가 없습니다.")
                return
            
            file_path, _ = QFileDialog.getSaveFileName(
                dialog, 
                "Excel 파일로 저장", 
                f"{log_type}_logs.xlsx", 
                "Excel 파일 (*.xlsx)"
            )
            
            if file_path:
                try:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = f"{log_type.upper()} 로그"
                    
                    # 헤더
                    headers = ["ID", "심각도", "생성 시간", "메시지"]
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                        ws.cell(row=1, column=col).font = Font(bold=True)
                    
                    # 로그 데이터
                    for row, entry in enumerate(log_entries, 2):
                        ws.cell(row=row, column=1, value=entry.get('Id', 'N/A'))
                        ws.cell(row=row, column=2, value=entry.get('Severity', 'N/A'))
                        ws.cell(row=row, column=3, value=format_time(entry.get('Created', 'N/A')))
                        ws.cell(row=row, column=4, value=entry.get('Message', 'N/A'))
                    
                    wb.save(file_path)
                    QMessageBox.information(dialog, "완료", f"로그가 {file_path}에 저장되었습니다.")
                except Exception as e:
                    QMessageBox.critical(dialog, "오류", f"Excel 저장 실패: {str(e)}")

        def clear_logs():
            confirm = QMessageBox.question(
                dialog,
                "확인",
                "모든 SEL 로그를 삭제하시겠습니까?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if confirm == QMessageBox.StandardButton.Yes:
                try:
                    server_manager.clear_sel_logs()
                    refresh_logs()
                    QMessageBox.information(dialog, "성공", "SEL 로그가 성공적으로 삭제되었습니다.")
                except Exception as e:
                    QMessageBox.critical(dialog, "오류", f"로그 삭제 실패: {str(e)}")

        # 이벤트 연결
        refresh_button.clicked.connect(refresh_logs)
        copy_button.clicked.connect(copy_logs_to_clipboard)
        excel_button.clicked.connect(export_logs_to_xlsx)
        
        if log_type == 'sel':
            clear_button.clicked.connect(clear_logs)
        
        severity_combo.currentTextChanged.connect(refresh_logs)
        search_input.textChanged.connect(refresh_logs)
        
        # 초기 로그 목록 로드
        refresh_logs()
        
        # 탭에 추가
        tab_widget.addTab(log_viewer_tab, "로그 뷰어")
        tab_widget.addTab(log_analysis_tab, "로그 분석")

        dialog.exec()

    except Exception as e:
        logger.error(f"{log_type.upper()} 로그 표시 중 오류 발생: {str(e)}")
        error_dialog = ErrorDialog(
            "로그 조회 오류",
            "로그를 표시하는 중 오류가 발생했습니다.",
            str(e),
            parent
        )
        error_dialog.exec()

def show_sel_log_popup(parent):
    show_log_popup(parent, 'sel')

def show_lc_log_popup(parent):
    show_log_popup(parent, 'lc')

def show_tsr_log_popup(parent):
    logger.debug("TSR 로그 수집 시도")
    main_window = parent.window()
    if not hasattr(main_window, 'server_section'):
        error_dialog = ErrorDialog(
            "서버 연결 오류",
            "서버가 연결되어 있지 않습니다.",
            "서버를 먼저 연결한 후 다시 시도해주세요.",
            parent
        )
        error_dialog.exec()
        return

    server_info = main_window.server_section.current_server_info
    if not server_info:
        error_dialog = ErrorDialog(
            "서버 연결 오류",
            "서버 정보를 찾을 수 없습니다.",
            "서버를 선택한 후 다시 시도해주세요.",
            parent
        )
        error_dialog.exec()
        return

    progress_dialog = QProgressDialog("TSR 로그 수집 중...", "취소", 0, 100, parent)
    progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
    progress_dialog.setWindowTitle("TSR 로그 수집")
    progress_dialog.setAutoClose(True)
    progress_dialog.setMinimumDuration(0)
    progress_dialog.show()

    try:
        server_manager = DellServerManager(
            ip=server_info['IP'],
            port=server_info['PORT'],
            auth=(server_info['USERNAME'], server_info['PASSWORD'])
        )

        def update_progress(progress):
            progress_dialog.setValue(int(progress))

        tsr_file = server_manager.collect_tsr_log(progress_callback=update_progress)
        
        if tsr_file:
            QMessageBox.information(parent, "완료", f"TSR 로그가 성공적으로 수집되었습니다.\n저장 위치: {tsr_file}")
        else:
            error_dialog = ErrorDialog(
                "TSR 로그 수집 오류",
                "TSR 로그 수집에 실패했습니다.",
                "서버 연결 상태를 확인하고 다시 시도해주세요.",
                parent
            )
            error_dialog.exec()

    except Exception as e:
        logger.error(f"TSR 로그 수집 중 오류 발생: {str(e)}")
        error_dialog = ErrorDialog(
            "TSR 로그 수집 오류",
            "TSR 로그 수집에 실패했습니다.",
            str(e),
            parent
        )
        error_dialog.exec()
    finally:
        progress_dialog.close()

def update_all_status():
    """모든 시스템 상태 정보 업데이트"""
    try:
        main_window = get_main_window()
        if not main_window or not hasattr(main_window, 'server_section'):
            return
            
        server_info = main_window.server_section.current_server_info
        if not server_info:
            return
            
        server_manager = main_window.server_section.server_manager
        if not server_manager:
            return
            
        # CPU 정보 업데이트
        cpu_info = server_manager.fetch_processors_info()
        if cpu_info:
            # CPU 상태 업데이트 로직
            pass
            
        # 메모리 정보 업데이트
        memory_info = server_manager.fetch_memory_info()
        if memory_info:
            # 메모리 상태 업데이트 로직
            pass
            
        # 스토리지 정보 업데이트
        storage_info = server_manager.fetch_storage_info()
        if storage_info:
            # 스토리지 상태 업데이트 로직
            pass
            
        # 전원 정보 업데이트
        power_info = server_manager.fetch_psu_info()
        if power_info:
            # 전원 상태 업데이트 로직
            pass
            
        logger.debug("시스템 상태 정보 업데이트 완료")
        
    except Exception as e:
        logger.error(f"시스템 상태 정보 업데이트 실패: {str(e)}")

def show_task_manager(parent):
    """작업 관리자 다이얼로그 표시"""
    logger.debug("작업 관리자 다이얼로그 표시 시도")
    
    main_window = parent.window()
    if not hasattr(main_window, 'server_section'):
        logger.warning("서버 섹션을 찾을 수 없음")
        error_dialog = ErrorDialog(
            "서버 연결 오류",
            "서버가 연결되어 있지 않습니다.",
            "서버를 먼저 연결한 후 다시 시도해주세요.",
            parent
        )
        error_dialog.exec()
        return

    server_info = main_window.server_section.current_server_info
    if not server_info:
        logger.warning("서버 정보가 없음")
        error_dialog = ErrorDialog(
            "서버 연결 오류",
            "서버 정보를 찾을 수 없습니다.",
            "서버를 선택한 후 다시 시도해주세요.",
            parent
        )
        error_dialog.exec()
        return

    try:
        server_manager = DellServerManager(
            ip=server_info['IP'],
            port=server_info['PORT'],
            auth=(server_info['USERNAME'], server_info['PASSWORD'])
        )

        dialog = QDialog(parent)
        dialog.setWindowTitle("작업 관리")
        dialog.resize(900, 600)
        layout = QVBoxLayout(dialog)

        # 필터 컨트롤 추가
        filter_layout = QHBoxLayout()
        status_combo = QComboBox()
        status_combo.addItems(['전체', 'Completed', 'Failed', 'Running'])
        search_input = QLineEdit()
        search_input.setPlaceholderText("작업 ID 또는 이름으로 검색")
        
        filter_layout.addWidget(QLabel("상태:"))
        filter_layout.addWidget(status_combo)
        filter_layout.addWidget(search_input)
        layout.addLayout(filter_layout)

        # 작업 목록 트리 위젯
        tree_widget = QTreeWidget(dialog)
        tree_widget.setHeaderLabels(["작업 ID", "작업 종류", "상태", "진행률", "시작 시간", "종료 시간"])
        tree_widget.setSelectionMode(QTreeWidget.SelectionMode.ExtendedSelection)
        
        # 컬럼 너비 최적화
        tree_widget.setColumnWidth(0, 150)  # 작업 ID
        tree_widget.setColumnWidth(1, 200)  # 작업 종류
        tree_widget.setColumnWidth(2, 100)  # 상태
        tree_widget.setColumnWidth(3, 80)   # 진행률
        tree_widget.setColumnWidth(4, 150)  # 시작 시간
        tree_widget.setColumnWidth(5, 150)  # 종료 시간
        layout.addWidget(tree_widget)

        # 버튼 레이아웃
        button_layout = QHBoxLayout()
        refresh_button = QPushButton("새로고침")
        refresh_button.setIcon(QIcon("refresh_icon.png"))
        delete_button = QPushButton("선택 작업 삭제")
        delete_button.setIcon(QIcon("delete_icon.png"))
        
        button_layout.addWidget(refresh_button)
        button_layout.addWidget(delete_button)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        def add_job_to_tree(job_details):
            item = QTreeWidgetItem(tree_widget)
            item.setText(0, job_details.get('Id', 'N/A'))
            item.setText(1, job_details.get('Name', 'N/A'))
            
            status = job_details.get('JobState', 'N/A')
            item.setText(2, status)
            
            # 상태에 따른 색상 설정
            if status == 'Completed':
                item.setForeground(2, QColor('green'))
            elif status == 'Failed':
                item.setForeground(2, QColor('red'))
            elif status == 'Running':
                item.setForeground(2, QColor('blue'))
            
            # 진행률을 프로그레스바로 표시
            progress = job_details.get('PercentComplete', 0)
            progress_bar = QProgressBar()
            progress_bar.setValue(progress)
            progress_bar.setStyleSheet("""
                QProgressBar {
                    border: 1px solid grey;
                    border-radius: 2px;
                    text-align: center;
                }
                QProgressBar::chunk {
                    background-color: #2196F3;
                }
            """)
            tree_widget.setItemWidget(item, 3, progress_bar)
            
            # 시간 형식 개선
            start_time = job_details.get('StartTime', 'N/A')
            end_time = job_details.get('EndTime', 'N/A')
            item.setText(4, format_time(start_time) if start_time != 'N/A' else 'N/A')
            item.setText(5, format_time(end_time) if end_time != 'N/A' else 'N/A')

        def refresh_jobs():
            tree_widget.clear()
            try:
                jobs = server_manager.fetch_job_queue()
                job_items = []
                has_running_jobs = False
                
                for job in jobs.get('Members', []):
                    job_id = job['@odata.id'].split('/')[-1]
                    job_details = server_manager.fetch_job_details(job_id)
                    
                    # 진행 중인 작업 확인
                    if job_details.get('JobState') == 'Running':
                        has_running_jobs = True
                    
                    # 필터링 적용
                    if status_combo.currentText() != '전체' and job_details.get('JobState') != status_combo.currentText():
                        continue
                    
                    search_text = search_input.text().lower()
                    if search_text and search_text not in job_details.get('Id', '').lower() and \
                       search_text not in job_details.get('Name', '').lower():
                        continue
                    
                    job_items.append((job_details.get('StartTime', ''), job_details))
                
                # 진행 중인 작업이 있으면 갱신 주기 변경
                if has_running_jobs:
                    timer.setInterval(5000)  # 5초마다 갱신
                else:
                    timer.setInterval(30000)  # 30초마다 갱신
                
                # 시작 시간 기준 내림차순 정렬
                job_items.sort(key=lambda x: x[0], reverse=True)
                
                for _, job_details in job_items:
                    add_job_to_tree(job_details)
                    
            except Exception as e:
                QMessageBox.critical(dialog, "오류", f"작업 목록 조회 실패: {str(e)}")

        def delete_selected_job():
            selected_items = tree_widget.selectedItems()
            if not selected_items:
                QMessageBox.warning(dialog, "경고", "삭제할 작업을 선택해주세요.")
                return
            
            job_ids = [item.text(0) for item in selected_items]
            confirm = QMessageBox.question(
                dialog, 
                "확인", 
                f"선택한 {len(job_ids)}개의 작업을 삭제하시겠습니까?\n" + \
                f"작업 ID: {', '.join(job_ids)}",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if confirm == QMessageBox.StandardButton.Yes:
                try:
                    for job_id in job_ids:
                        server_manager.delete_job(job_id)
                    refresh_jobs()
                    QMessageBox.information(dialog, "완료", f"{len(job_ids)}개의 작업이 삭제되었습니다.")
                except Exception as e:
                    QMessageBox.critical(dialog, "오류", f"작업 삭제 실패: {str(e)}")

        # 다이얼로그가 닫힐 때 타이머 정지를 위한 이벤트 처리
        def on_dialog_finished():
            timer.stop()
            logger.debug("작업 관리자 다이얼로그 종료: 타이머 정지")
            
        # 이벤트 연결
        refresh_button.clicked.connect(refresh_jobs)
        delete_button.clicked.connect(delete_selected_job)
        status_combo.currentTextChanged.connect(refresh_jobs)
        search_input.textChanged.connect(refresh_jobs)

        # 타이머 설정 및 시작
        timer = QTimer(dialog)
        timer.timeout.connect(refresh_jobs)
        dialog.finished.connect(on_dialog_finished)
        
        # 초기 작업 목록 로드 및 타이머 시작
        refresh_jobs()  # 다이얼로그가 열릴 때 첫 조회
        timer.start(30000)  # 30초 주기로 시작
        
        dialog.exec()

    except Exception as e:
        logger.error(f"작업 관리자 표시 중 오류 발생: {str(e)}")
        error_dialog = ErrorDialog(
            "작업 관리자 오류",
            "작업 관리자를 표시하는 중 오류가 발생했습니다.",
            str(e),
            parent
        )
        error_dialog.exec()

def format_time(time_str):
    """시간 형식을 보기 좋게 변환"""
    if time_str and time_str != 'N/A':
        try:
            date_parts = time_str.split('T')
            if len(date_parts) == 2:
                return f"{date_parts[0]} {date_parts[1][:8]}"
        except:
            pass
    return time_str
